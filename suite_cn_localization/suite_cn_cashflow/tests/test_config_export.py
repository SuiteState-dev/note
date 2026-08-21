# -*- coding: utf-8 -*-
"""R25-T3 — P-05 可移交性: cash-flow config export / import round-trip.

Acceptance (design §2 P-05): 导出 → (清空配置模拟换库) → 导入 → 配置与导出前 identical.
Plus the Q2 ruling: an item present in the DB but NOT in the imported file must be
KEPT *and* REPORTED (never silently kept, never silently deleted).
"""
import base64
import io

from odoo.tests import TransactionCase, tagged


@tagged('post_install', '-at_install')
class TestCashflowConfigExport(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        item = cls.env['cash.flow.item']
        cls.s_recv = item.search([('code', '=', 'S2')], limit=1)  # inflow
        cls.s_pay = item.search([('code', '=', 'S6')], limit=1)   # outflow
        cls.acc = cls.env['account.account'].create({
            'name': 'RT Other', 'code': 'ZZRT1', 'account_type': 'asset_current',
            'cn_default_cash_flow_item_debit_id': cls.s_pay.id,
            'cn_default_cash_flow_item_credit_id': cls.s_recv.id,
        })

    def _wizard(self):
        return self.env['suite.cn.cashflow.config'].create(
            {'company_id': self.company.id})

    def test_round_trip_identical(self):
        # -- export --
        w = self._wizard()
        w.action_export()
        self.assertTrue(w.export_file, "export produced no file")
        blob = base64.b64decode(w.export_file)

        # the produced workbook has both named sheets and our account row
        import openpyxl  # noqa: PLC0415
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        self.assertIn('现金流量项目', wb.sheetnames)
        self.assertIn('科目默认映射', wb.sheetnames)
        map_codes = [r[0] for r in wb['科目默认映射'].iter_rows(min_row=2, values_only=True)]
        self.assertIn('ZZRT1', map_codes, "mapped account missing from export")
        wb.close()

        # -- simulate 换库: wipe the mapping --
        self.acc.write({
            'cn_default_cash_flow_item_debit_id': False,
            'cn_default_cash_flow_item_credit_id': False,
        })
        self.assertFalse(self.acc.cn_default_cash_flow_item_debit_id)

        # -- import the same file back --
        w2 = self._wizard()
        w2.write({'import_file': base64.b64encode(blob),
                  'import_filename': 'cfg.xlsx'})
        w2.action_import()

        # -- identical to before the wipe --
        self.assertEqual(self.acc.cn_default_cash_flow_item_debit_id, self.s_pay,
                         "debit default not restored by import")
        self.assertEqual(self.acc.cn_default_cash_flow_item_credit_id, self.s_recv,
                         "credit default not restored by import")

    def test_leftover_reported(self):
        """Q2: a DB item not in the file is kept AND listed in the result."""
        w = self._wizard()
        w.action_export()
        blob = base64.b64decode(w.export_file)

        # a NEW account default that the exported file does not know about
        extra = self.env['account.account'].create({
            'name': 'RT Extra', 'code': 'ZZRT2', 'account_type': 'asset_current',
            'cn_default_cash_flow_item_debit_id': self.s_pay.id,
        })
        w2 = self._wizard()
        w2.write({'import_file': base64.b64encode(blob),
                  'import_filename': 'cfg.xlsx'})
        w2.action_import()

        # kept (incremental — not mirror-deleted)…
        self.assertEqual(extra.cn_default_cash_flow_item_debit_id, self.s_pay,
                         "incremental import must not clear a DB-only mapping")
        # …and reported so a human can decide (not silent).
        self.assertIn('ZZRT2', w2.result or '',
                      "a DB-only mapping must be listed in the import result, not "
                      "silently kept (§7.0 铁律1: 检查跑了 ≠ 有人看见)")
