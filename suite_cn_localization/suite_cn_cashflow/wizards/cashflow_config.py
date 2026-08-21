# -*- coding: utf-8 -*-
"""P-05 可移交性 — 现金流量分类配置的导出 / 导入通道 (R25-T3).

`cn_cash_flow_item_id` is the one accounting dimension this suite adds that has no
export channel of its own (design §2 P-05, R22-T5 盘点). Without one, a client who
wants to leave Odoo or change implementer cannot take the *configuration* with them
— we would have manufactured exactly the lock-in the 金蝶 material档 §6.7 records
(231 科目 mappings + templates that "no other system can eat once exported"). This
wizard closes that: it round-trips the CONFIG (statutory items + per-account default
mappings) through a plain, human-readable XLSX that any consultant — or any other
system — can read, and reads the same file back for a库-change / disaster restore.

Scope (design §2 P-05):
  * cash.flow.item master (编码/名称/类别/准则/序号)                 — sheet 1
  * account default mapping (借/贷 default cash flow item per account) — sheet 2
Line-level overrides (account.move.line.cn_cash_flow_item_id) are NOT here by
design: that field is a standard m2o, already an optional column on the journal-item
list, so Odoo's native list export takes it away; and it travels WITH the accounting
data on any DB migration — it is 账务数据, not a config asset (T3.3 结论, design §2).

Design rulings baked in (R25):
  * ONE workbook, TWO sheets — the consumer is the incoming implementer / the client
    themselves, not a tax-bureau importer; one file is easier for a human (Q1).
  * Reverse import is INCREMENTAL upsert (never a mirror-delete), BUT it must REPORT
    the "in DB, not in file" leftovers so a human decides — silent-keep is as wrong as
    silent-delete (Q2, the same §7.0 铁律1 as T1: 检查跑了 ≠ 有人看见).
"""
import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Human-readable Chinese written into / read back from the XLSX. We deliberately do
# NOT dump the raw selection key (operating / asbe) — the file is for humans and
# foreign systems (P-05). Import is lenient: it accepts the Chinese label, the English
# selection label, or the raw key, so a hand-edited file still reads back.
_CATEGORY_ZH = {'operating': '经营活动', 'investing': '投资活动', 'financing': '筹资活动'}
_STANDARD_ZH = {'asbe': '企业会计准则(ASBE)', 'assbe': '小企业会计准则(ASSBE)'}

_SHEET_ITEMS = '现金流量项目'
_SHEET_MAP = '科目默认映射'
_ITEM_COLS = ['编码', '名称', '类别', '准则', '序号']
_MAP_COLS = ['科目编码', '科目名称', '借方项目编码', '借方项目名称',
             '贷方项目编码', '贷方项目名称']


def _zh_to_key(value, mapping):
    """Map a human label (Chinese / English / raw key) back to a selection key.

    Matching is EXACT on: the raw key, the full label, the Chinese prefix before
    '(', or the English code inside '(...)'. A substring test would be wrong —
    「企业会计准则」 is a substring of 「小企业会计准则」, so ASSBE would mis-map to
    ASBE (caught by round-trip 造数复跑, R25-T3)."""
    if not value:
        return False
    v = str(value).strip()
    vl = v.lower()
    v_prefix = v.split('(')[0].strip()
    for key, zh in mapping.items():
        zh_prefix = zh.split('(')[0].strip()
        paren = zh.split('(')[1].rstrip(')').lower() if '(' in zh else ''
        if vl == key or v == zh or v_prefix == zh_prefix or (paren and vl == paren):
            return key
    return False


class CashflowConfigWizard(models.TransientModel):
    _name = 'suite.cn.cashflow.config'
    _description = 'Cash Flow Config Export / Import (P-05 可移交性)'

    company_id = fields.Many2one(
        'res.company', string='Company', required=True,
        default=lambda self: self.env.company,
        help="账户编码按公司取值(company-dependent)，导出/导入以此公司口径。")

    # -- export side --
    export_file = fields.Binary(string='Export File', readonly=True, attachment=False)
    export_filename = fields.Char(readonly=True)

    # -- import side --
    import_file = fields.Binary(string='Import File')
    import_filename = fields.Char()
    result = fields.Text(string='Result', readonly=True)

    # ------------------------------------------------------------------ export
    def _cf_items(self):
        """Statutory items visible to this company = global (company_id empty) +
        this company's own. active_test=False so an archived-but-referenced item
        (retired via archive, R12-A1) still travels with the config."""
        return self.env['cash.flow.item'].with_context(active_test=False).search(
            ['|', ('company_id', '=', False), ('company_id', '=', self.company_id.id)],
            order='sequence, code')

    def _cf_mapped_accounts(self):
        """Accounts carrying a default on either direction, in this company's
        code口径 (code is company-dependent)."""
        Acc = self.env['account.account'].with_company(self.company_id)
        return Acc.search(
            ['|', ('cn_default_cash_flow_item_debit_id', '!=', False),
                  ('cn_default_cash_flow_item_credit_id', '!=', False)],
            order='code')

    def action_export(self):
        self.ensure_one()
        import xlsxwriter  # noqa: PLC0415  (Odoo-standard late import)
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True})
        hdr = wb.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1})
        cell = wb.add_format({'border': 1})

        ws1 = wb.add_worksheet(_SHEET_ITEMS)
        for c, h in enumerate(_ITEM_COLS):
            ws1.write(0, c, h, hdr)
        ws1.set_column(0, 1, 22)
        ws1.set_column(2, 3, 18)
        for r, it in enumerate(self._cf_items(), start=1):
            ws1.write(r, 0, it.code or '', cell)
            ws1.write(r, 1, it.name or '', cell)
            ws1.write(r, 2, _CATEGORY_ZH.get(it.category, ''), cell)
            ws1.write(r, 3, _STANDARD_ZH.get(it.standard, ''), cell)
            ws1.write_number(r, 4, it.sequence or 0, cell)

        ws2 = wb.add_worksheet(_SHEET_MAP)
        for c, h in enumerate(_MAP_COLS):
            ws2.write(0, c, h, hdr)
        ws2.set_column(0, 0, 14)
        ws2.set_column(1, 1, 30)
        ws2.set_column(2, 5, 18)
        for r, acc in enumerate(self._cf_mapped_accounts(), start=1):
            d = acc.cn_default_cash_flow_item_debit_id
            k = acc.cn_default_cash_flow_item_credit_id
            ws2.write(r, 0, acc.code or '', cell)
            ws2.write(r, 1, acc.name or '', cell)
            ws2.write(r, 2, d.code or '', cell)
            ws2.write(r, 3, d.name or '', cell)
            ws2.write(r, 4, k.code or '', cell)
            ws2.write(r, 5, k.name or '', cell)
        wb.close()
        buf.seek(0)
        fname = '现金流量配置_%s.xlsx' % (self.company_id.name or '')
        self.write({
            'export_file': base64.b64encode(buf.read()),
            'export_filename': fname,
        })
        # Reload the same wizard so the download link (the binary field) is shown —
        # a one-file, in-UI download, no developer mode needed (§7.0 铁律1 可达性).
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ------------------------------------------------------------------ import
    def _read_sheets(self):
        """Parse the uploaded workbook into {sheet_name: [row_dicts]} keyed by the
        Chinese header row. openpyxl is bundled (base_import uses it); read-only."""
        import openpyxl  # noqa: PLC0415
        try:
            data = base64.b64decode(self.import_file)
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001 — surface a friendly UserError
            raise UserError(_("无法读取该文件，请确认是本向导导出的 .xlsx：%s", e))
        out = {}
        for name in (_SHEET_ITEMS, _SHEET_MAP):
            if name not in wb.sheetnames:
                out[name] = []
                continue
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                out[name] = []
                continue
            header = [str(h).strip() if h is not None else '' for h in rows[0]]
            recs = []
            for raw in rows[1:]:
                rec = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
                if any(v not in (None, '') for v in rec.values()):
                    recs.append(rec)
            out[name] = recs
        wb.close()
        return out

    def _find_item(self, code):
        code = (str(code).strip() if code is not None else '')
        if not code:
            return self.env['cash.flow.item']
        return self.env['cash.flow.item'].with_context(active_test=False).search(
            ['&', ('code', '=', code),
             '|', ('company_id', '=', False), ('company_id', '=', self.company_id.id)],
            limit=1)

    def action_import(self):
        self.ensure_one()
        if not self.import_file:
            raise UserError(_("请先上传要导入的配置文件。"))
        sheets = self._read_sheets()
        Item = self.env['cash.flow.item']

        # -- sheet 1: cash.flow.item upsert (incremental — never delete) --
        created = updated = 0
        seen_item_codes = set()
        for rec in sheets.get(_SHEET_ITEMS, []):
            code = (str(rec.get('编码') or '').strip())
            if not code:
                continue
            seen_item_codes.add(code)
            vals = {
                'name': (str(rec.get('名称') or '').strip()) or code,
                'category': _zh_to_key(rec.get('类别'), _CATEGORY_ZH) or 'operating',
                'standard': _zh_to_key(rec.get('准则'), _STANDARD_ZH),
                'sequence': int(rec.get('序号') or 10),
            }
            item = self._find_item(code)
            if item:
                item.write(vals)
                updated += 1
            else:
                # A new code from the file is created as a GLOBAL statutory item
                # (company_id empty) — the shipped set is global, and this keeps a
                # clean-DB restore aligned with the data file. Company-specific
                # custom items re-import as global; documented limitation (README).
                Item.create(dict(vals, code=code))
                created += 1

        # -- sheet 2: per-account default mapping (incremental) --
        mapped = 0
        skipped_accounts = []
        seen_account_codes = set()
        Acc = self.env['account.account'].with_company(self.company_id)
        for rec in sheets.get(_SHEET_MAP, []):
            acode = (str(rec.get('科目编码') or '').strip())
            if not acode:
                continue
            seen_account_codes.add(acode)
            account = Acc.search([('code', '=', acode)], limit=1)
            if not account:
                skipped_accounts.append(acode)
                continue
            debit = self._find_item(rec.get('借方项目编码'))
            credit = self._find_item(rec.get('贷方项目编码'))
            account.write({
                'cn_default_cash_flow_item_debit_id': debit.id or False,
                'cn_default_cash_flow_item_credit_id': credit.id or False,
            })
            mapped += 1

        # -- Q2: report what is in the DB but NOT in the file (kept, not touched) --
        db_items = self._cf_items()
        leftover_items = db_items.filtered(lambda i: i.code not in seen_item_codes)
        db_accounts = self._cf_mapped_accounts()
        leftover_accounts = db_accounts.filtered(
            lambda a: a.code not in seen_account_codes)

        lines = [_("导入完成（公司：%s）。", self.company_id.name)]
        lines.append(_("现金流量项目：新建 %(c)s 项，更新 %(u)s 项。", c=created, u=updated))
        lines.append(_("科目默认映射：更新 %(m)s 个科目。", m=mapped))
        if skipped_accounts:
            lines.append(_(
                "⚠ 文件中有 %(n)s 个科目编码在本公司科目表中不存在，已跳过（未新建科目）：\n  %(l)s",
                n=len(skipped_accounts), l='、'.join(skipped_accounts)))
        if leftover_items:
            lines.append(_(
                "⚠ 库中存在、本次文件未覆盖的现金流量项目 %(n)s 项（已保留，未改动，请自行确认是否需要）：",
                n=len(leftover_items)))
            lines += ['  [%s] %s' % (i.code, i.name) for i in leftover_items]
        if leftover_accounts:
            lines.append(_(
                "⚠ 库中存在、本次文件未覆盖的科目默认映射 %(n)s 个（已保留，未改动）：",
                n=len(leftover_accounts)))
            lines += ['  [%s] %s' % (a.code, a.name) for a in leftover_accounts]
        if not leftover_items and not leftover_accounts and not skipped_accounts:
            lines.append(_("库与文件一致，无遗留项。"))
        self.write({'result': '\n'.join(lines)})
        _logger.info("suite_cn_cashflow config import: +%d/~%d items, %d accounts, "
                     "%d leftover items, %d leftover maps",
                     created, updated, mapped, len(leftover_items),
                     len(leftover_accounts))
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
