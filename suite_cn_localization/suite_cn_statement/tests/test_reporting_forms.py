# -*- coding: utf-8 -*-
"""R23-T4-1 / R23-T2 durable acceptance guards.

These institutionalise two coverage gaps R23 exposed on the live UI that every
prior round's programmatic acceptance missed:

* **Row-number ordering (T4-1)** — R21 only checked row-number *completeness*
  (no missing / no duplicate rows), never *order*. Completeness and order are
  independent: the official export shipped all 33 rows in fully scrambled order.
  The Chinese export orders rows by ``suite.cn.statement.row.sequence`` and splits
  the Balance Sheet into two side-by-side columns; each column's printed 行次 must
  be strictly increasing on its own. This test asserts exactly that, at the form
  level (``export_to_cn_xlsx`` renders ``form.row_ids.sorted('sequence')`` split by
  section, so the form ordering *is* the export ordering — no accounting data
  needed, which keeps the test portable to any DB).

* **No dangling aggregation term (T2)** — a generated ``bal_begin`` / ``ytd``
  expression referencing a ``CODE.label`` that does not exist raises
  "Could not expand term" the moment the report is opened, invisible to any
  ``_get_lines`` smoke test that does not hit that line. This re-runs the
  install-time self-check and fails if any dangling term ships.
"""
from odoo import fields
from odoo.addons.suite_cn_statement.hooks import _log_dangling_expressions
from odoo.exceptions import UserError
from odoo.tests import TransactionCase, tagged


@tagged('post_install', '-at_install')
class TestReportingForms(TransactionCase):

    def _row_numbers(self, rows):
        """Non-empty integer 行次 of data (non-header) rows, in render order."""
        out = []
        for row in rows:
            if row.is_header or not (row.row_no or '').strip():
                continue
            # Every printed row number must be an integer (a stray label such as
            # '1.00' or a non-numeric string is itself a defect worth catching).
            out.append(int(row.row_no.strip()))
        return out

    def _assert_strictly_increasing(self, numbers, label):
        self.assertTrue(numbers, "%s: expected at least one numbered row" % label)
        for a, b in zip(numbers, numbers[1:]):
            self.assertLess(
                a, b,
                "%s: 行次 not strictly increasing — %s then %s (order defect)"
                % (label, a, b))
            # R24-T2: 报送 row numbers must also be CONTIGUOUS (no gap). The tax
            # bureau importer positions rows by 行次, so a gap is a mis-placement,
            # not a cosmetic ordering issue — design §7.1 T4-1 and background §8
            # both promise "无缺号"; assert it rather than let the docs over-claim
            # a check the code never performed. ASSBE三表 and ASBE BS (左1-34 /
            # 右35-72) are all contiguous, so this passes today. If a future
            # statutory form legitimately skips numbers (common on 附表), add a
            # form-level exemption flag — never delete this assertion.
            self.assertEqual(
                b, a + 1,
                "%s: 行次 gap — %s then %s (missing %s)" % (label, a, b, a + 1))

    def test_row_no_ordered(self):
        """T4-1: each export column's 行次 is strictly increasing (BS two sides
        checked independently — the two-column layout is the easiest to scramble)."""
        forms = self.env['suite.cn.statement.form'].search([])
        self.assertTrue(forms, "no reporting forms installed")
        for form in forms:
            # R26-T2: an 年报 form borrows its rows from the 月季报 sibling
            # (rows_from_id) — assert on the EFFECTIVE rows so a borrowing form is
            # still ordering-checked (and a form that borrows nothing still uses its
            # own rows).
            rows = form._effective_rows().sorted('sequence')
            if not form.has_row_no:
                # A no-row-number form (tax-bureau ASBE P&L) must carry NO row
                # numbers at all — assert every row's row_no is empty. Without
                # this branch a form that should have dropped its 行次 could pass
                # this test simply by keeping stale numbers; here the ABSENCE is
                # the invariant (R24-T2).
                leftover = [
                    r.row_no for r in rows if (r.row_no or '').strip()]
                self.assertFalse(
                    leftover,
                    "%s: has_row_no=False but row_no present on %d row(s): %s"
                    % (form.report_id.name, len(leftover), leftover))
                continue
            if form.layout == 'two_column':
                for section in ('left', 'right'):
                    nums = self._row_numbers(
                        [r for r in rows if r.section == section])
                    self._assert_strictly_increasing(
                        nums, "%s [%s]" % (form.report_id.name, section))
            else:
                self._assert_strictly_increasing(
                    self._row_numbers(rows), form.report_id.name)

    def _asbe_form(self, kind):
        """The ASBE Balance-Sheet ('bs') or P&L ('pl') form, or None if absent."""
        ref = 'l10n_cn_reports.l10n_cn_asbe_%s' % kind
        rep = self.env.ref(ref, raise_if_not_found=False)
        if not rep:
            return None
        return self.env['suite.cn.statement.form'].search(
            [('report_id', '=', rep.id)], limit=1)

    def test_crossfoot_evaluator(self):
        """R24 review: 数值对 + 行次对 ≠ 表内勾稽对. The evaluator must catch a
        subtotal that does not equal the sum of its shown detail rows — that break
        equals the balance of a dropped 官方有·报送无 row (买入返售 / 以前年度损益调整)
        and would be rejected by the tax-bureau importer. Tested on synthetic render
        data so it is portable (no accounting data needed)."""
        bs = self._asbe_form('bs')
        if not bs:
            self.skipTest('l10n_cn_asbe_bs not installed')
        rep = bs.report_id

        # An all-zero render is trivially consistent (every 合计 = Σ 0 = 0).
        def row(no, v, section):
            return {'row_no': str(no), 'name': str(no), 'is_header': False,
                    'values': [v], 'section': section}
        left = [row(i, 0.0, 'left') for i in range(1, 35)]    # rows 1..34
        right = [row(i, 0.0, 'right') for i in range(35, 73)]  # rows 35..72
        data_ok = {'ncols': 1, 'layout': 'two_column', 'left': left, 'right': right}
        self.assertFalse(
            rep._cn_crossfoot_breaks(bs, data_ok),
            "consistent (all-zero) data reported a spurious cross-foot break")

        # Break row 14 by 7000 — the exact signature of a dropped 买入返售 balance
        # the official 流动资产合计 includes but no shown detail row carries.
        left[13] = row(14, 7000.0, 'left')  # index 13 == 行次 14
        data_bad = {'ncols': 1, 'layout': 'two_column', 'left': left, 'right': right}
        breaks = rep._cn_crossfoot_breaks(bs, data_bad)
        self.assertTrue(breaks, "cross-foot break not detected")
        hit = [b for b in breaks if b[1] == 0 and b[0].startswith('14 ')]
        self.assertTrue(hit, "expected a break on subtotal row 14, got %s" % breaks)
        self.assertAlmostEqual(hit[0][4], 7000.0, places=2,
                               msg="break delta must equal the dropped-row balance")

    def test_crossfoot_live(self):
        """EVERY shipped form that has a cross-foot rule set (ASSBE BS/PL, ASBE
        BS/PL, CF — §12.3 判据4) must cross-foot on the live report with clean demo
        data. This is what keeps the guard's coverage honest against §12.3: a form
        with rules that does NOT reconcile here means either a mis-encoded rule or a
        dropped row carrying a balance — both must be caught in CI, not on upload.
        Companies with a dropped-row balance would legitimately WARN, so we assert
        on the clean steady state of the CN demo companies.

        SCOPE (R25, do not misread): this zero-break assertion is a promise about the
        TEST FIXTURE, NOT about a customer database. In a real ledger unclassified cash
        flow (cn_cf_s_unc != 0) is the NORMAL state -- 漏标 happens daily -- so the CF
        form will legitimately break 20 = 7+13+19. That is a DATA-QUALITY signal (kind=
        'unclassified' -> "去补标后重新导出"), NOT a code defect. Do not "fix" the code
        when a customer lib shows a CF cross-foot break; classify the flow instead."""
        forms = self.env['suite.cn.statement.form'].search([])
        companies = self.env['res.company'].search(
            [('chart_template', 'in', ('cn', 'cn_large_bis'))])
        checked = 0
        for form in forms:
            rep = form.report_id
            if not rep._cn_crossfoot_rules(form):
                continue  # no rule set registered for this form
            for company in companies:
                report = rep.with_company(company).with_context(
                    allowed_company_ids=[company.id])
                data = report._cn_prepare(form, report.get_options({}))
                self.assertFalse(
                    data['crossfoot_breaks'],
                    "%s [%s]: unexpected cross-foot break(s) on clean data — a "
                    "mis-encoded rule or a dropped row with a balance: %s"
                    % (form.title, company.name, data['crossfoot_breaks']))
                checked += 1
        self.assertTrue(checked, "no forms with cross-foot rules were checked")

    def test_crossfoot_messages_two_states(self):
        """R25-T1: the in-file warning line (`_cn_crossfoot_messages`) must appear
        with the right content when a break exists and be ABSENT when clean — the
        accountant clicking Export reads the file, never the server log (§7.0 铁律1).
        R24 shipped the renderer but only tested the math; assert the visible line."""
        bs = self._asbe_form('bs')
        if not bs:
            self.skipTest('l10n_cn_asbe_bs not installed')
        rep = bs.report_id

        # clean → no line at all (no noise on a good file)
        data_ok = {'crossfoot_breaks': [], 'col_headers': ['期末余额'],
                   'col_group_labels': []}
        self.assertEqual(rep._cn_crossfoot_messages(data_ok), [],
                         "a clean form must emit NO in-file warning line")

        # one break → a header line + one detail line naming the row and delta
        data_bad = {
            'crossfoot_breaks': [('14 = Σ(1..13) 流动资产合计', 0, 7000.0, 0.0, 7000.0)],
            'col_headers': ['期末余额']}
        msgs = rep._cn_crossfoot_messages(data_bad)
        self.assertTrue(msgs, "a break must emit an in-file warning")
        self.assertIn('1 处', msgs[0], "header should state the break count")
        body = '\n'.join(msgs)
        self.assertIn('14 ', body, "the broken row number must be named")
        self.assertIn('7000', body.replace(',', ''),
                      "the delta (points at the dropped row) must be shown")

    def test_crossfoot_message_kinds(self):
        """R25: a 未分类现金流量 break (kind='unclassified') must be worded as a
        fixable data-quality gap ('去补标', 'not a report error') — NOT like a
        structural 买入返售/以前年度损益调整 break ('本表可能不适用'). Sharing the
        wording would make the accountant think the CF report is broken (§15.3)."""
        bs = self._asbe_form('bs')
        if not bs:
            self.skipTest('l10n_cn_asbe_bs not installed')
        rep = bs.report_id

        structural = {'crossfoot_breaks': [
            ('14 = 流动资产合计', 0, 7000.0, 0.0, 7000.0, 'structural')],
            'col_headers': ['期末余额']}
        smsg = '\n'.join(rep._cn_crossfoot_messages(structural))
        self.assertIn('本表可能不适用', smsg)
        self.assertNotIn('未分类', smsg)

        unclassified = {'crossfoot_breaks': [
            ('20 = 现金净增加额', 0, 300.0, 0.0, 300.0, 'unclassified')],
            'col_headers': ['本月']}
        umsg = '\n'.join(rep._cn_crossfoot_messages(unclassified))
        self.assertIn('未分类', umsg)
        self.assertIn('补标', umsg)
        self.assertIn('非报表错误', umsg)
        self.assertNotIn('本表可能不适用', umsg)
        self.assertNotIn('选错报送主体', umsg)

    def test_crossfoot_inequality(self):
        """R30-P2: 行9 ≥ 行10+11+12+13 (存货 ≥ 其中项之和) is an INEQUALITY (op='gte',
        source 财会〔2011〕17号 附录 编制说明 3.4). A SHORTFALL — 存货 < Σ其中项, i.e. the
        sub-items exceed their parent (a mapping/data error) — MUST break with an
        inequality-specific message. An EXCESS — 存货 > Σ其中项, the normal case since
        其中项 are non-exhaustive (半成品/产成品/… may be unlisted) — MUST NOT break.
        Synthetic render data (portable, no accounting data needed)."""
        rep = self.env.ref('l10n_cn_reports.l10n_cn_assbe_bs', raise_if_not_found=False)
        if not rep:
            self.skipTest('l10n_cn_assbe_bs not installed')
        form = self.env['suite.cn.statement.form'].search(
            [('report_id', '=', rep.id)], limit=1)
        if not form:
            self.skipTest('ASSBE BS form not present')

        def row(no, v):
            # ASSBE BS is two-column: 1..30 资产 (left), 31..53 负债权益 (right).
            section = 'left' if no <= 30 else 'right'
            return {'row_no': str(no), 'name': str(no), 'is_header': False,
                    'values': [v], 'section': section}

        def render(row9, sub):
            vals = {9: row9, 10: sub[0], 11: sub[1], 12: sub[2], 13: sub[3]}
            left = [row(i, vals.get(i, 0.0)) for i in range(1, 31)]
            right = [row(i, 0.0) for i in range(31, 54)]
            return {'ncols': 1, 'layout': 'two_column', 'left': left, 'right': right}

        # SHORTFALL: 存货 100 < Σ其中项 120 → break, delta = −20, op='gte'
        breaks = rep._cn_crossfoot_breaks(form, render(100.0, (40, 40, 40, 0)))
        hit = [b for b in breaks if b[0].startswith('9 ') and b[1] == 0]
        self.assertTrue(hit, "存货 < Σ其中项 must break 行9 inequality; got %s" % breaks)
        b = hit[0]
        self.assertEqual(b[6], 'gte', "the 行9 break must carry op='gte'")
        self.assertAlmostEqual(b[4], -20.0, places=2,
                               msg="delta must equal the shortfall (−20)")
        msg = '\n'.join(rep._cn_crossfoot_messages(
            {'crossfoot_breaks': [b], 'col_headers': ['期末余额']}))
        self.assertIn('内含项', msg, "inequality message must name 内含项")
        self.assertIn('超出', msg, "the direction word (超出) must show")
        self.assertIn('20', msg.replace(',', ''), "the shortfall amount must show")
        self.assertNotIn('本表可能不适用', msg,
                         "inequality must NOT reuse the structural (换报表) wording")
        # R30-P2 审稿 ①②: the outbound (税局-facing) line must NOT self-incriminate
        # ('此为数据错误') nor make the absolute (待核 Q2b) claim ('子项不应大于总额').
        self.assertNotIn('数据错误', msg, "outbound text must not self-incriminate to 税局")
        self.assertNotIn('不应', msg, "outbound text must not make the absolute-subset claim")

        # EXCESS: 存货 200 > Σ其中项 120 → the normal non-exhaustive case, NO 行9 break
        breaks2 = rep._cn_crossfoot_breaks(form, render(200.0, (40, 40, 40, 0)))
        hit2 = [b for b in breaks2 if b[0].startswith('9 ')]
        self.assertFalse(hit2, "存货 > Σ其中项 (non-exhaustive) must NOT break; got %s" % hit2)

    def test_crossfoot_nonneg_inventory(self):
        """R37-T2 (B-69): 行9 存货 ≥ 0 is a lone-row SEMANTIC assertion (op='nonneg',
        🔴 自推·无原文). B-69 drives 存货 negative while 资产 = 负债权益 stays self-
        consistent (−6000 asset exactly cancels −6000 equity), so every balance-class
        rule passes — only this assertion catches it. A NEGATIVE 存货 must break; a
        zero/positive 存货 must NOT. The message must say 存货为负, must NOT reuse the
        内含项超出 (gte) or 换报表 (structural) wording, and must NOT self-incriminate
        to the 税局 (数据错误). Both准则 (ASSBE / ASBE) BS carry the rule."""
        for report_xmlid, split, maxrow in (
                ('l10n_cn_reports.l10n_cn_assbe_bs', 30, 53),
                ('l10n_cn_reports.l10n_cn_asbe_bs', 34, 72)):
            rep = self.env.ref(report_xmlid, raise_if_not_found=False)
            if not rep:
                continue
            form = self.env['suite.cn.statement.form'].search(
                [('report_id', '=', rep.id)], limit=1)
            if not form:
                continue

            def render(inv):
                rows = [{'row_no': str(i), 'name': str(i), 'is_header': False,
                         'values': [inv if i == 9 else 0.0],
                         'section': 'left' if i <= split else 'right'}
                        for i in range(1, maxrow + 1)]
                return {'ncols': 1, 'layout': 'two_column',
                        'left': [r for r in rows if r['section'] == 'left'],
                        'right': [r for r in rows if r['section'] == 'right']}

            def nonneg_hits(inv):
                return [b for b in rep._cn_crossfoot_breaks(form, render(inv))
                        if b[0].startswith('9 ') and len(b) > 6 and b[6] == 'nonneg']

            hit = nonneg_hits(-6000.0)
            self.assertTrue(
                hit, "%s: 存货<0 must break nonneg; got %s"
                % (report_xmlid, rep._cn_crossfoot_breaks(form, render(-6000.0))))
            self.assertAlmostEqual(hit[0][2], -6000.0, places=2,
                                   msg="nonneg actual must be the 存货 value")
            self.assertFalse(nonneg_hits(0.0), "%s: 存货=0 must NOT break" % report_xmlid)
            self.assertFalse(nonneg_hits(6000.0), "%s: 存货>0 must NOT break" % report_xmlid)

            msg = '\n'.join(rep._cn_crossfoot_messages(
                {'crossfoot_breaks': [hit[0]], 'col_headers': ['期末余额']}))
            self.assertIn('存货', msg, "message must name 存货")
            self.assertIn('负', msg, "message must say the value is negative")
            self.assertIn('6000', msg.replace(',', ''), "the negative amount must show")
            self.assertNotIn('内含项', msg, "must NOT reuse the gte (内含项超出) wording")
            self.assertNotIn('本表可能不适用', msg, "must NOT reuse structural (换报表) wording")
            self.assertNotIn('数据错误', msg, "outbound text must not self-incriminate to 税局")

    def test_crossfoot_inequality_tolerance(self):
        """R30-P2 Q3: the 'gte' comparison is SINGLE-SIDED but shares eq's 0.01
        tolerance (delta = round(actual−expect, 2); break iff delta < −0.01). A
        sub-cent / one-cent shortfall must NOT break (rounding noise, not a real
        sub-item-exceeds-parent error); a two-cent+ shortfall must break."""
        rep = self.env.ref('l10n_cn_reports.l10n_cn_assbe_bs', raise_if_not_found=False)
        if not rep:
            self.skipTest('l10n_cn_assbe_bs not installed')
        form = self.env['suite.cn.statement.form'].search(
            [('report_id', '=', rep.id)], limit=1)
        if not form:
            self.skipTest('ASSBE BS form not present')

        def render(row9):
            # Σ其中项 (rows 10..13) fixed at 100.00; vary 存货 (row 9) around it.
            vals = {9: row9, 10: 100.0, 11: 0.0, 12: 0.0, 13: 0.0}
            left = [{'row_no': str(i), 'name': str(i), 'is_header': False,
                     'values': [vals.get(i, 0.0)], 'section': 'left'}
                    for i in range(1, 31)]
            right = [{'row_no': str(i), 'name': str(i), 'is_header': False,
                      'values': [0.0], 'section': 'right'} for i in range(31, 54)]
            return {'ncols': 1, 'layout': 'two_column', 'left': left, 'right': right}

        def broke(row9):
            return bool([b for b in rep._cn_crossfoot_breaks(form, render(row9))
                         if b[0].startswith('9 ')])

        # expect − 0.005 (sub-cent) → NO break
        self.assertFalse(broke(99.995), "0.005 shortfall must be within tolerance")
        # expect − 0.01 (one cent) → NO break (tolerance is ±0.01 inclusive)
        self.assertFalse(broke(99.99), "0.01 shortfall must be within tolerance")
        # expect − 0.02 → break (past tolerance)
        self.assertTrue(broke(99.98), "0.02 shortfall must break the inequality")
        # An EXCESS of any size (single-sided) must never break.
        self.assertFalse(broke(100.50), "an excess must never break (single-sided)")

    # ------------------------------------------------ R30-T2a 往来行方向分流 (B-58)
    def _assbe_bs_live(self):
        """(rep, form, report, company, A, post) for a live ASSBE-BS direction test,
        or None (caller skips). ``A(code)`` → account, ``post(acc, amt, date)`` posts a
        balanced misc move (amt>0 debit / <0 credit) and returns it."""
        rep = self.env.ref('l10n_cn_reports.l10n_cn_assbe_bs', raise_if_not_found=False)
        if not rep:
            return None
        form = self.env['suite.cn.statement.form'].search(
            [('report_id', '=', rep.id)], limit=1)
        company = self.env['res.company'].search([('chart_template', '=', 'cn')], limit=1)
        if not form or not company:
            return None
        A = self.env['account.account'].with_company(company)

        def acc(code):
            return A.search([('code', '=', code)], limit=1)

        journal = self.env['account.journal'].search(
            [('company_id', '=', company.id), ('type', '=', 'general')], limit=1)
        ctr = acc('4001') or A.search([('account_type', '=', 'equity')], limit=1)
        if not (journal and ctr and acc('1122') and acc('2203')):
            return None

        def post(a, amt, date):
            d, c = (amt, 0) if amt > 0 else (0, -amt)
            mv = self.env['account.move'].with_company(company).create({
                'move_type': 'entry', 'journal_id': journal.id, 'date': date,
                'line_ids': [(0, 0, {'account_id': a.id, 'debit': d, 'credit': c}),
                             (0, 0, {'account_id': ctr.id, 'debit': c, 'credit': d})]})
            mv._post(soft=False)
            return mv

        report = rep.with_company(company).with_context(allowed_company_ids=[company.id])
        return rep, form, report, company, acc, post

    def _dir_render(self, report, form):
        opts = report.get_options({'date': {
            'mode': 'range', 'filter': 'custom',
            'date_from': '2026-01-01', 'date_to': '2026-12-31'}})
        data = report._cn_prepare(form, opts)
        val = {}
        for r in (data.get('left', []) + data.get('right', [])):
            rn = (r.get('row_no') or '').strip()
            if rn.isdigit():
                val[int(rn)] = r['values']
        return val, data['crossfoot_breaks']

    @staticmethod
    def _balance_breaks(breaks):
        """R37-T2: crossfoot breaks EXCLUDING the 存货≥0 (op='nonneg') assertion. The
        _assbe_bs_live fixture posts every entry against counterpart 4001 (= 生产成本 =
        在产品/存货 under the R36 override口径), so 存货 goes NEGATIVE by the posted amount —
        a FIXTURE ARTIFACT, not a balance/dir-split defect. These tests assert the
        balance-class rules (53=30, subtotal cascades) stay green; the independent,
        fixture-induced nonneg 存货 signal is filtered out here (it is exercised on its
        own in test_crossfoot_nonneg_inventory)."""
        return [b for b in breaks if not (len(b) > 6 and b[6] == 'nonneg')]

    def test_dir_split_double_column(self):
        """R30-T2a go/no-go: 期末 and 年初 columns must EACH judge direction on their
        OWN net. Fixture 年初 1122 借 500 / 期末 1122 贷 200 → 应收 [期末 0, 年初 500],
        预收 [期末 200, 年初 0]. A single whole-report evaluation applied to all columns
        would mis-render the 年初 column — this asserts four cells, not two."""
        ctx = self._assbe_bs_live()
        if not ctx:
            self.skipTest('ASSBE BS live prerequisites missing')
        rep, form, report, company, acc, post = ctx
        cols = [(c.column_group, c.expression_label) for c in form.column_ids.sorted('sequence')]
        # test assumes column order [期末 balance, 年初 bal_begin]
        self.assertEqual(cols[:2], [('primary', 'balance'), ('primary', 'bal_begin')],
                         "unexpected column order: %s" % cols)
        post(acc('1122'), 500, '2025-06-01')    # opening (年初) debit 500
        post(acc('1122'), -700, '2026-06-01')   # → 期末 net credit 200
        val, breaks = self._dir_render(report, form)
        self.assertAlmostEqual(val[4][0], 0.0, 2, "期末: net credit → 应收 0")
        self.assertAlmostEqual(val[4][1], 500.0, 2, "年初: net debit → 应收 500")
        self.assertAlmostEqual(val[34][0], 200.0, 2, "期末: net credit → 预收 200 (positive)")
        self.assertAlmostEqual(val[34][1], 0.0, 2, "年初: net debit → 预收 0")
        self.assertFalse(self._balance_breaks(breaks),
                         "direction split must not break crossfoot: %s" % breaks)

    def test_dir_split_cross_and_balance(self):
        """R30-T2a: four 往来 科目 non-zero & direction-crossed (2203/2202 non-zero is
        demo-invisible). 1122借500 / 2203借80 / 1123贷30 / 2202借40 →
        应收 580 (=pos1122+pos2203) / 预收 0 / 预付 40 (=pos2202) / 应付 30 (=neg1123).
        53=30 must hold (split moves asset→liability by equal amounts) and _CN_CROSSFOOT
        must stay green (subtotal delta-cascade)."""
        ctx = self._assbe_bs_live()
        if not ctx:
            self.skipTest('ASSBE BS live prerequisites missing')
        rep, form, report, company, acc, post = ctx
        if not (acc('1123') and acc('2202')):
            self.skipTest('1123/2202 missing')
        post(acc('1122'), 500, '2026-06-01')
        post(acc('2203'), 80, '2026-06-01')
        post(acc('1123'), -30, '2026-06-01')
        post(acc('2202'), 40, '2026-06-01')
        val, breaks = self._dir_render(report, form)
        self.assertAlmostEqual(val[4][0], 580.0, 2, "应收 = pos(1122)+pos(2203) = 500+80")
        self.assertAlmostEqual(val[34][0], 0.0, 2, "预收 = 0")
        self.assertAlmostEqual(val[5][0], 40.0, 2, "预付 = pos(1123)+pos(2202) = 0+40")
        self.assertAlmostEqual(val[33][0], 30.0, 2, "应付 = neg(2202)+neg(1123) = 0+30")
        # 53 = 30 (balance preserved by the split)
        self.assertAlmostEqual(val[30][0], val[53][0], 2,
                               "53=30 must hold after the split: %s vs %s" % (val[30][0], val[53][0]))
        self.assertFalse(self._balance_breaks(breaks),
                         "_CN_CROSSFOOT must stay green after the split: %s" % breaks)

    def test_accum_deprec_positive(self):
        """R30-P3 ①: 减：累计折旧 (行19) is filled as a POSITIVE magnitude on the 报送
        form (财会〔2011〕17号 编制说明 (13): 累计折旧 科目 贷方余额 = 累计折旧额), and
        行20 账面价值 = 行18 − 行19. Odoo renders the contra-asset 累计折旧 as −3000, so the
        form must present +3000 and the crossfoot rule subtracts. Fixture 原价 10000 /
        折旧 3000 → 行18 10000, 行19 +3000, 行20 7000, crossfoot green."""
        ctx = self._assbe_bs_live()
        if not ctx:
            self.skipTest('ASSBE BS live prerequisites missing')
        rep, form, report, company, acc, post = ctx
        a1601, a1602 = acc('1601'), acc('1602')
        exp = self.env['account.account'].with_company(company).search(
            [('account_type', '=', 'expense')], limit=1)
        if not (a1601 and a1602 and exp):
            self.skipTest('1601/1602/expense account missing')
        post(a1601, 10000, '2026-06-01')          # 固定资产原价 debit 10000
        # 累计折旧: credit 1602 (contra-asset), debit an expense — post directly
        mv = self.env['account.move'].with_company(company).create({
            'move_type': 'entry', 'journal_id':
                self.env['account.journal'].search(
                    [('company_id', '=', company.id), ('type', '=', 'general')], limit=1).id,
            'date': '2026-06-01',
            'line_ids': [(0, 0, {'account_id': a1602.id, 'debit': 0, 'credit': 3000}),
                         (0, 0, {'account_id': exp.id, 'debit': 3000, 'credit': 0})]})
        mv._post(soft=False)
        val, breaks = self._dir_render(report, form)
        self.assertAlmostEqual(val[18][0], 10000.0, 2, "行18 固定资产原价")
        self.assertAlmostEqual(val[19][0], 3000.0, 2,
                               "行19 减：累计折旧 must render POSITIVE 3000, not −3000")
        self.assertAlmostEqual(val[20][0], 7000.0, 2, "行20 账面价值 = 18−19 = 7000")
        self.assertFalse(self._balance_breaks(breaks),
                         "20=18−19 must hold with a positive 行19: %s" % breaks)

    # ------------------------------------------ R30-T2b 往来明细级 D/C 方向分流
    def _asbe_bs_dc_live(self):
        """(rep, form, report, company, acc, post, render) for a live ASBE-BS D/C test,
        or None (caller skips)."""
        rep = self.env.ref('l10n_cn_reports.l10n_cn_asbe_bs', raise_if_not_found=False)
        if not rep:
            return None
        form = self.env['suite.cn.statement.form'].search(
            [('report_id', '=', rep.id)], limit=1)
        company = self.env['res.company'].search(
            [('chart_template', '=', 'cn_large_bis')], limit=1)
        if not form or not company:
            return None
        A = self.env['account.account'].with_company(company)

        def acc(code):
            return A.search([('code', '=', code)], limit=1)

        journal = self.env['account.journal'].search(
            [('company_id', '=', company.id), ('type', '=', 'general')], limit=1)
        ctr = A.search([('account_type', '=', 'equity')], limit=1)
        if not (journal and ctr and acc('2202') and acc('1123') and acc('1231.03')):
            return None

        def post(a, amt, date='2026-06-01'):
            d, c = (amt, 0) if amt > 0 else (0, -amt)
            self.env['account.move'].with_company(company).create({
                'move_type': 'entry', 'journal_id': journal.id, 'date': date,
                'line_ids': [(0, 0, {'account_id': a.id, 'debit': d, 'credit': c}),
                             (0, 0, {'account_id': ctr.id, 'debit': c, 'credit': d})]}
            )._post(soft=False)

        report = rep.with_company(company).with_context(allowed_company_ids=[company.id])

        def render():
            self.env.invalidate_all()   # B-61: engine must see fresh accounts
            opts = report.get_options({'date': {
                'mode': 'range', 'filter': 'custom',
                'date_from': '2026-01-01', 'date_to': '2026-12-31'}})
            data = report._cn_prepare(form, opts)
            val = {}
            for r in (data.get('left', []) + data.get('right', [])):
                rn = (r.get('row_no') or '').strip()
                if rn.isdigit():
                    val[int(rn)] = r['values']
            return val, data['crossfoot_breaks'], data.get('dc_notice') or []

        return rep, form, report, company, A, acc, post, render

    def test_dc_split_subaccounts_is_the_proof(self):
        """R30-T2b — THE proof (惯例6): 明细级 D/C only manifests WITH 子账户. Fixture:
        2202.01 贷500 / 2202.02 借100 / 1123.01 借300 / 1123.02 贷80 / 1231.03 贷30 →
        应付账款(行39) = −(2202C+1123C) = −(−500−80) = 580;
        预付款项(行7)  = 2202D+1123D+1231.03 = 100+300−30 = 370 (1231.03 归宿=−30, 取值).
        53=30 preserved, crossfoot green, NO 科目级 notice (明细级 reached)."""
        ctx = self._asbe_bs_dc_live()
        if not ctx:
            self.skipTest('ASBE BS D/C live prerequisites missing')
        rep, form, report, company, A, acc, post, render = ctx
        # build 往来 子账户 (mixed direction)
        s = {}
        for code, atype in [('2202.01', 'liability_payable'), ('2202.02', 'liability_payable'),
                            ('1123.01', 'asset_prepayments'), ('1123.02', 'asset_prepayments')]:
            s[code] = A.create({'code': code, 'name': 'DC ' + code, 'account_type': atype})
        post(s['2202.01'], -500)
        post(s['2202.02'], 100)
        post(s['1123.01'], 300)
        post(s['1123.02'], -80)
        post(acc('1231.03'), -30)                 # 坏账准备(预付) 贷方
        val, breaks, notice = render()
        self.assertAlmostEqual(val[39][0], 580.0, 2,
                               "应付账款 = −(2202C+1123C) = 580, got %s" % val[39][0])
        self.assertAlmostEqual(val[7][0], 370.0, 2,
                               "预付款项 = 2202D+1123D+1231.03 = 370, got %s" % val[7][0])
        # 53=30 (asset total == liab+equity total)
        self.assertAlmostEqual(val[34][0], val[72][0], 2,
                               "34=72 must hold after the split: %s vs %s" % (val[34][0], val[72][0]))
        self.assertFalse(breaks, "_CN_CROSSFOOT must stay green: %s" % breaks)
        self.assertFalse(notice, "with 子账户 the split is 明细级 → NO 科目级 notice")

    def test_dc_no_subaccounts_triggers_notice_not_break(self):
        """R30-T2b go/no-go: WITHOUT 子账户 the D/C degrades to 科目级 → a 口径 NOTICE
        (NOT a crossfoot break, NOT a raise). 🔴 The rendered values are IDENTICAL to
        pre-change (Case A: 2202C=net, 2202D=0) — this零差异 is the EXPECTED result, NOT
        proof of correctness (惯例6: the proof is test_dc_split_subaccounts_is_the_proof).
        This test only asserts the notice fires and nothing breaks/raises."""
        ctx = self._asbe_bs_dc_live()
        if not ctx:
            self.skipTest('ASBE BS D/C live prerequisites missing')
        rep, form, report, company, A, acc, post, render = ctx
        post(acc('2202'), -400)                   # bare 2202, no 子账户
        post(acc('1123'), 220)
        post(acc('1231.03'), -30)
        val, breaks, notice = render()
        # values equal the un-split net (应付=400, 预付=220−30=190) — EXPECTED, NOT PROOF
        self.assertAlmostEqual(val[39][0], 400.0, 2, "Case A 应付=科目级 net (预期非证据)")
        self.assertAlmostEqual(val[7][0], 190.0, 2, "Case A 预付=科目级 net (预期非证据)")
        self.assertTrue(notice, "无子账户 must raise the 科目级口径 notice")
        self.assertIn('科目级', notice[0])
        self.assertNotIn('数据错误', notice[0])   # 口径声明, not a break/error
        self.assertFalse(breaks, "科目级 notice must NOT be a crossfoot break")

    # ------------------------------------------------------------ R26-T2 model
    def _form(self, xmlid):
        return self.env.ref('suite_cn_statement.%s' % xmlid,
                            raise_if_not_found=False) or \
            self.env.ref('suite_cn_cashflow_statement.%s' % xmlid,
                         raise_if_not_found=False)

    def test_period_scope_assignments(self):
        """R26-T2: the three-dim enum is populated as ruled — ASSBE BS is 'any'
        (月季=年报 cell-identical), ASSBE PL/CF are monthly_quarterly with an annual
        sibling, ASBE BS/PL are monthly_quarterly + executed with NO annual (🔴 no
        material). A regression here is the silent-口径-swap risk."""
        bs = self._form('cn_form_bs')
        if not bs:
            self.skipTest('ASSBE forms not installed')
        self.assertEqual(bs.period_scope, 'any')
        self.assertEqual(bs.standard_version, 'na')
        self.assertEqual(self._form('cn_form_pl').period_scope, 'monthly_quarterly')
        self.assertEqual(self._form('cn_form_pl_annual').period_scope, 'annual')
        self.assertEqual(self._form('cn_form_pl_annual').rows_from_id,
                         self._form('cn_form_pl'),
                         '年报 must borrow rows from 月季报 (single source of truth)')
        asbe_bs = self._form('cn_form_asbe_bs')
        if asbe_bs:
            self.assertEqual(asbe_bs.period_scope, 'monthly_quarterly',
                             'ASBE 手上是季报版; annual has no material → must NOT be any')
            self.assertEqual(asbe_bs.standard_version, 'executed')

    def test_form_selection_and_no_silent_fallback(self):
        """🔴 R26-T2: exact period wins; 'any' serves both cadences; a report with a
        monthly form but NO annual form must resolve to EMPTY and the export must
        raise an explicit UserError — NEVER silently fall back to the monthly form
        (that would ship 季报 口径 under a 年报 filing)."""
        pl = self._form('cn_form_pl')
        if not pl:
            self.skipTest('ASSBE forms not installed')
        rep_assbe_pl = pl.report_id
        # exact + annual both resolve to the right form
        self.assertEqual(
            rep_assbe_pl._cn_statement_form('monthly_quarterly'), pl)
        self.assertEqual(
            rep_assbe_pl._cn_statement_form('annual'),
            self._form('cn_form_pl_annual'))
        # ASSBE BS 'any' serves an annual request (no annual-specific form exists)
        rep_bs = self._form('cn_form_bs').report_id
        self.assertEqual(rep_bs._cn_statement_form('annual'), self._form('cn_form_bs'))
        # ASBE PL: annual unsupported → EMPTY resolution, explicit raise
        asbe_pl = self._form('cn_form_asbe_pl')
        if asbe_pl:
            rep_asbe_pl = asbe_pl.report_id
            self.assertFalse(
                rep_asbe_pl._cn_statement_form('annual'),
                'ASBE annual must resolve EMPTY, not fall back to the 季报 form')
            with self.assertRaises(UserError):
                rep_asbe_pl._cn_resolve_form_or_raise('annual')

    def test_export_wizard_action_has_views(self):
        """🔴 R27 regression: the 中式版式 button in the account.report bar is
        dispatched through dispatch_report_action → client actionService.doAction
        DIRECTLY, bypassing web call_button/clean_action. So the framework never
        补全 view_mode→views; _preprocessAction's act_window branch hard-reads
        action.views.map(...), and a missing `views` key is a client-side
        undefined.map TypeError (向导整个打不开). Python tests can't run the JS
        preprocessing, so guard the contract here: the returned act_window dict
        MUST carry an explicit non-empty `views`."""
        pl = self._form('cn_form_pl')
        if not pl:
            self.skipTest('ASSBE forms not installed')
        act = pl.report_id.action_open_cn_export_wizard({'report_id': pl.report_id.id})
        self.assertEqual(act.get('type'), 'ir.actions.act_window')
        self.assertTrue(act.get('views'),
                        'act_window dispatched via account_reports must ship an '
                        'explicit `views` (view_mode is not enough — no clean_action)')
        self.assertTrue(any(v[1] == 'form' for v in act['views']))

    def test_annual_requires_full_fiscal_year(self):
        """🔴 R26 review: an 年报 form over a NON-full-year period (e.g. Jan–Jun) is a
        mislabeled 半年报 — 本年累计半年 | 上年同期半年 under the 年报 表名/表号. Same
        silent-wrong-口径 class as the no-fallback rule, on the period axis. Must
        hard-stop; a monthly form at any period must NOT (guard is annual-only."""
        from datetime import date
        pl_annual = self._form('cn_form_pl_annual')
        company = self._cn_company()
        if not pl_annual or not company:
            self.skipTest('ASSBE forms / CN company not installed')
        rep = pl_annual.report_id
        fy = company.compute_fiscalyear_dates(date(2026, 6, 30))
        comp = [{'id': company.id}]
        full = {'date': {'date_from': str(fy['date_from']),
                         'date_to': str(fy['date_to'])}, 'companies': comp}
        partial = {'date': {'date_from': str(fy['date_from']),
                            'date_to': str(fy['date_from'])}, 'companies': comp}
        # full fiscal year → OK
        rep._cn_assert_annual_period(pl_annual, full)
        # partial period under an 年报 form → explicit stop
        with self.assertRaises(UserError):
            rep._cn_assert_annual_period(pl_annual, partial)
        # a monthly form is exempt at any period
        rep._cn_assert_annual_period(self._form('cn_form_pl'), partial)

    # ------------------------------------------------------------ R26-T1 layout
    def _cn_company(self):
        return self.env['res.company'].search(
            [('chart_template', 'in', ('cn', 'cn_large_bis'))], limit=1)

    def _render_load(self, form, period_scope='monthly_quarterly'):
        """Render a form to XLSX and return an openpyxl worksheet for cell asserts."""
        import io
        from datetime import date
        from openpyxl import load_workbook
        company = self._cn_company()
        rep = form.report_id.with_company(company).with_context(
            allowed_company_ids=[company.id])
        options = rep.get_options({})
        if period_scope == 'annual':
            # An 年报 form hard-requires a FULL fiscal year (see
            # _cn_assert_annual_period, else UserError) — anchor options to the whole
            # FY so the annual export renders instead of raising the 半年报 guard.
            anchor = fields.Date.to_date(
                (options.get('date') or {}).get('date_to')) or date(2026, 12, 31)
            fy = company.compute_fiscalyear_dates(anchor)
            options = {**options, 'date': {
                **(options.get('date') or {}),
                'mode': 'range',
                'date_from': str(fy['date_from']),
                'date_to': str(fy['date_to']),
            }}
        result = rep.export_to_cn_xlsx(options, period_scope)
        wb = load_workbook(io.BytesIO(result['file_content']), data_only=True)
        return wb.active

    def test_template_geometry_single(self):
        """R26-T1: the corrected tax-bureau geometry on a single-column form —
        blank column A, 项目 BEFORE 行次, header on row 5 (1-indexed), metadata rows.
        Asserts OUR renderer against the spec (the official .xls fixture pins the
        exact metadata columns in test_template_alignment, pending upload)."""
        pl = self._form('cn_form_pl')
        if not pl or not self._cn_company():
            self.skipTest('ASSBE forms / CN company not installed')
        ws = self._render_load(pl)
        # blank column A on the header row
        self.assertIn(ws.cell(row=5, column=1).value, (None, ''),
                      'column A must be blank (whole block shifts one column right)')
        self.assertEqual(ws.cell(row=5, column=2).value, '项目',
                         '项目 must be the first data column (B)')
        self.assertEqual(ws.cell(row=5, column=3).value, '行次',
                         '行次 must come AFTER 项目 (R26-T1)')
        # title + 适用说明 on row 1, starting at column B (blank column A, R27)
        self.assertIn(ws.cell(row=1, column=1).value, (None, ''),
                      'column A must be blank on the title row too (R27)')
        self.assertIn('利润表', str(ws.cell(row=1, column=2).value or ''))
        # metadata labels present in the header block
        block = '\n'.join(
            str(ws.cell(row=r, column=c).value or '')
            for r in (2, 3, 4) for c in range(1, 8))
        for label in ('纳税人识别号', '纳税人名称', '所属期起', '所属期止'):
            self.assertIn(label, block, '%s metadata cell missing' % label)

    def test_template_geometry_two_column(self):
        """R26-T1: the Balance-Sheet two-column geometry — blank column A, left head
        资 产 at B, right head 负债和所有者权益 immediately after the left value block
        with NO gap column (8 data columns B..I for a 2-value BS)."""
        bs = self._form('cn_form_bs')
        if not bs or not self._cn_company():
            self.skipTest('ASSBE forms / CN company not installed')
        ws = self._render_load(bs, 'monthly_quarterly')
        self.assertIn(ws.cell(row=5, column=1).value, (None, ''),
                      'column A must be blank')
        self.assertEqual(ws.cell(row=5, column=2).value, '资产')
        self.assertEqual(ws.cell(row=5, column=3).value, '行次')
        # ncol=2 → left block B..E (2..5), right block starts at F (6): no gap column
        self.assertEqual(ws.cell(row=5, column=6).value, '负债和所有者权益',
                         'right side must start at column F — no gap column (R26-T1)')
        self.assertEqual(ws.cell(row=5, column=7).value, '行次')

    def test_template_alignment(self):
        """R26-T1 / R27 逐格对位: our export must equal the OFFICIAL tax-bureau
        template cell-for-cell on the header block (rows 0–4). Expected values are
        PARSED from the official .xls fixtures — never transcribed (the prose spec was
        wrong twice: it missed the blank column A and the bare metadata labels). The
        fixture IS the expected value. SkipTest until the fixtures are uploaded to
        tests/fixtures/ (design §15.4 / R26-T1)."""
        import os
        try:
            import xlrd
        except ImportError:
            self.skipTest('xlrd not available to parse .xls fixtures')
        fx = os.path.join(os.path.dirname(__file__), 'fixtures')
        files = {
            'monthly_quarterly':
                '财务报表报送与信息采集（小企业会计准则）月季报.xls',
            'annual': '财务报表报送与信息采集（小企业会计准则）年报.xls',
        }
        if not os.path.exists(os.path.join(fx, files['monthly_quarterly'])):
            self.skipTest('official .xls template fixtures not yet uploaded '
                          '(tests/fixtures/) — see design §15.4 / R26-T1')
        if not self._cn_company():
            self.skipTest('no CN company installed')
        # Official sheet name → our form xmlid. The Balance Sheet is period-identical
        # (period_scope='any') so both files map to the same cn_form_bs.
        sheet_form = {
            '资产负债表': 'cn_form_bs',
            '利润表_月季报': 'cn_form_pl', '利润表_年': 'cn_form_pl_annual',
            '现金流量表_月季报': 'cn_form_cf', '现金流量表_年': 'cn_form_cf_annual',
        }
        HEADER_ROWS = 5  # official rows 0..4 (0-indexed): 表名/表号/识别号/所属期/列头
        checked = 0
        for scope, fname in files.items():
            path = os.path.join(fx, fname)
            if not os.path.exists(path):
                continue
            book = xlrd.open_workbook(path)
            for sheet in book.sheets():
                xmlid = sheet_form.get(sheet.name)
                self.assertTrue(
                    xmlid, 'official sheet %r has no mapped form' % sheet.name)
                form = self._form(xmlid)
                if not form:
                    self.skipTest('form %s not installed' % xmlid)
                ws = self._render_load(form, scope)
                for r in range(min(HEADER_ROWS, sheet.nrows)):
                    for c in range(sheet.ncols):
                        exp = sheet.cell_value(r, c)
                        # Assert ONLY where the official template carries text. Its
                        # value cells (识别号/名称/期间/金额) are blank — those are ours
                        # to pre-fill and must not be pinned to the empty template.
                        if exp is None or (isinstance(exp, str)
                                           and not exp.strip()):
                            continue
                        got = ws.cell(row=r + 1, column=c + 1).value
                        # Compare as text; internal spacing (表号 的 3 空格) is
                        # significant, so strip only the outer whitespace.
                        self.assertEqual(
                            ('' if got is None else str(got)).strip(),
                            str(exp).strip(),
                            '%s [%s] cell (row=%d,col=%d): official %r != export %r'
                            % (sheet.name, scope, r, c, exp, got))
                        checked += 1
        # Guard against a silently-empty parse (fixtures present but unreadable would
        # otherwise pass vacuously — the "检查跑了≠有人看见" 铁律).
        self.assertGreater(
            checked, 40,
            'expected many header cells to be asserted; parsed too few (%d)' % checked)

    def test_row_no_rule_on_official_fixture(self):
        """🔴 R27-T1-3: pin the 行次 NUMBERING RULE ITSELF against the OFFICIAL
        template, not just against our own render. The rule (induced from two
        independent statutory samples across both准则 — §4.5.8): 只给数据行编号；
        节标题行行次留空；左栏从 1 起，右栏接续（NOT restart at 1）。

        test_row_no_ordered already checks our forms are monotone — but that only
        proves internal consistency. This parses the ASSBE 官方 BS .xls (53 数据行 +
        5 节标题) and asserts the RULE holds in the ground truth, which promotes the
        rule from `观察` to `verified` (design §15.4) and lets the ASBE 未执行版
        推演行次 be a 'verified-rule extrapolation' rather than 凭空推演.

        This is orthogonal to the alignment test (which pins header rows 0–4 text/
        geometry) and to test_row_no_ordered (which pins OUR render) — it pins the
        official DATA-region numbering scheme."""
        import os
        try:
            import xlrd
        except ImportError:
            self.skipTest('xlrd not available to parse .xls fixtures')
        fx = os.path.join(os.path.dirname(__file__), 'fixtures')
        fname = '财务报表报送与信息采集（小企业会计准则）月季报.xls'
        if not os.path.exists(os.path.join(fx, fname)):
            self.skipTest('official .xls template fixtures not yet uploaded')
        book = xlrd.open_workbook(os.path.join(fx, fname))
        sheet = next((s for s in book.sheets() if s.name == '资产负债表'), None)
        self.assertTrue(sheet, 'official 资产负债表 sheet missing from fixture')
        HEADER_ROWS = 5  # rows 0–4 are 表名/表号/识别号/所属期/列头; data starts at 5

        def _side(label_col, num_col):
            """Return the list of printed 行次 down one column, asserting every
            section-header row (label ends in ：) carries a BLANK 行次 and every
            numbered row carries an integer — i.e. 只给数据行编号 / 标题留空."""
            nums = []
            for r in range(HEADER_ROWS, sheet.nrows):
                label = str(sheet.cell_value(r, label_col)).strip()
                raw = sheet.cell_value(r, num_col)
                if not label:
                    continue  # pure spacer row
                if label.endswith(('：', ':')):
                    self.assertIn(
                        raw, ('', None),
                        '节标题 %r must have BLANK 行次, got %r' % (label, raw))
                    continue
                self.assertNotIn(
                    raw, ('', None),
                    'data row %r must carry a 行次' % label)
                nums.append(int(raw))
            return nums

        def _contiguous_from(nums, first, label):
            self.assertTrue(nums, '%s: no numbered rows parsed' % label)
            self.assertEqual(nums[0], first,
                             '%s: 行次 must start at %d, got %d'
                             % (label, first, nums[0]))
            self.assertEqual(
                nums, list(range(first, first + len(nums))),
                '%s: 行次 must be strictly increasing & contiguous — got %s'
                % (label, nums))

        left = _side(1, 2)   # 资产 label col 1, 行次 col 2
        right = _side(5, 6)  # 负债和所有者权益 label col 5, 行次 col 6
        # 左栏从 1 起
        _contiguous_from(left, 1, 'BS left (资产)')
        # 右栏接续：first right number == last left number + 1 (NOT a restart at 1)
        _contiguous_from(right, left[-1] + 1, 'BS right (负债和所有者权益)')
        # Pin the concrete ASSBE sample shape (§4.5.11: 左1–30 / 右31–53) so a
        # future fixture swap that silently changes the scheme is caught.
        self.assertEqual((left[0], left[-1]), (1, 30), 'ASSBE BS 左栏应为 1–30')
        self.assertEqual((right[0], right[-1]), (31, 53), 'ASSBE BS 右栏应为 31–53')

    def test_bs_deliberate_deviations_from_bureau_typos(self):
        """🔴 R45-T3:资产负债表【故意偏离税局表样】—— PIN 两处已登记偏离(税局模板自身笔误)。

        税局采集件【自身】有两处笔误,我方 form 按【准则正确写法】故意不照搬(Safi R45-T3 拍板;
        §4.5.7.3「行名逐字一致」加例外条款)。本用例把两处偏离【钉住】,防下一轮有人为「对齐夹具」
        改回错别字(夹具与 form 抄同一个错时普通对位会全绿、缺陷静默过关——这正是 PIN 的价值)。
        夹具原件【不动】(税局件忠实副本、含其笔误)。
          行次24 生产性生物资产:税局【月季报】误作「生物性生物资产」(生物×2),【年报】正确
            「生产性生物资产」——同一采集件两册在此行本就不一致;我方取正确写法(=年报)。
          行次30 资产总计:税局【月季报+年报两册】此行均作「资产合计」,但其右栏行53作
            「负债和所有者权益…总计」——左合计/右总计,税局模板自身不一致;我方两侧对称作
            「总计」(行30 资产总计 / 行53 …总计),与准则正表左右对称一致。"""
        import os
        try:
            import xlrd
        except ImportError:
            self.skipTest('xlrd not available to parse .xls fixtures')
        fx = os.path.join(os.path.dirname(__file__), 'fixtures')
        files = {
            'monthly_quarterly':
                '财务报表报送与信息采集（小企业会计准则）月季报.xls',
            'annual': '财务报表报送与信息采集（小企业会计准则）年报.xls',
        }
        if not os.path.exists(os.path.join(fx, files['monthly_quarterly'])):
            self.skipTest('official .xls template fixtures not yet uploaded')

        def _bs_name(fname, row_no):
            """原件资产负债表某行次的行名(左栏 label col1 / 行次 col2)。"""
            book = xlrd.open_workbook(os.path.join(fx, fname))
            sh = next(s for s in book.sheets() if s.name == '资产负债表')
            for r in range(5, sh.nrows):
                if sh.cell_value(r, 2) == float(row_no):
                    return str(sh.cell_value(r, 1)).strip()
            return None

        form = self._form('cn_form_bs')
        if not form:
            self.skipTest('cn_form_bs not installed')

        def _our(row_no):
            rows = form.row_ids.filtered(
                lambda r: (r.row_no or '').strip() == str(row_no))
            self.assertEqual(len(rows), 1, 'cn_form_bs 应恰有一条 行次%s' % row_no)
            return rows.name

        # ── 行次24:月季夹具=税局笔误 / 年报夹具=正确 / 我方=正确(=年报、≠月季)。
        self.assertEqual(_bs_name(files['monthly_quarterly'], 24), '生物性生物资产',
                         '月季报夹具行24应为税局原样笔误「生物性生物资产」(夹具不动)')
        self.assertEqual(_bs_name(files['annual'], 24), '生产性生物资产',
                         '年报夹具行24应为正确「生产性生物资产」')
        self.assertEqual(_our(24), '生产性生物资产',
                         '🔴 我方行24须按准则作「生产性生物资产」,不得回退月季报笔误')
        self.assertNotEqual(_our(24), _bs_name(files['monthly_quarterly'], 24),
                            '我方行24与月季报夹具【故意不逐字一致】(勿"对齐修复")')
        # ── 行次30:两册夹具均=税局笔误「资产合计」/ 我方=对称「资产总计」(≠两册夹具)。
        self.assertEqual(_bs_name(files['monthly_quarterly'], 30), '资产合计',
                         '月季报夹具行30应为税局原样「资产合计」(夹具不动)')
        self.assertEqual(_bs_name(files['annual'], 30), '资产合计',
                         '年报夹具行30应为税局原样「资产合计」(两册同,夹具不动)')
        self.assertEqual(_our(30), '资产总计',
                         '🔴 R45-T3:我方行30作对称「资产总计」(右栏行53亦为…总计),'
                         '不得为「对齐夹具」改回「资产合计」')
        self.assertNotEqual(_our(30), _bs_name(files['monthly_quarterly'], 30),
                            '我方行30与夹具【故意不逐字一致】(第二处已登记偏离)')

    def test_no_dangling_expressions(self):
        """T2: no generated aggregation expression references a missing term.

        🔴 R46 两半(否则空绿):原用例只 assertFalse(dangling) —— 打桩 _log_dangling_expressions
        →[] 时天然全绿,无法区分「查过、无悬空」与「检测器空转」(R46 实测:该函数在
        our_exprs 为空时亦提前 return [])。① 正向哨兵:注入一条【已知悬空】聚合表达式,断言
        检测器确能抓到(检测器空转 → 红);② 撤下探针后,真数据上无悬空。"""
        Expr = self.env['account.report.expression']
        IMD = self.env['ir.model.data']
        # 取本模块一条聚合表达式,借它的报表定位到我方报表。
        imd0 = IMD.search([('module', '=', 'suite_cn_statement'),
                           ('model', '=', 'account.report.expression')], limit=1)
        self.assertTrue(imd0, '前置:本模块须有聚合表达式(否则整个自检无对象=空绿根因)')
        report = Expr.browse(imd0.res_id).report_line_id.report_id
        # ① 正向哨兵:在【新建的干净报表行】(无 groupby)上挂一条引用不存在 term 的聚合表达式,
        #    须被检出。挂 imd(module=本模块)让检测器视其为「本模块表达式」纳入扫描。
        probe_line = self.env['account.report.line'].create({
            'name': 'R46 Probe Line', 'report_id': report.id})
        probe = Expr.create({
            'report_line_id': probe_line.id, 'label': 'r46_probe',
            'engine': 'aggregation', 'formula': 'ZZ_NO_SUCH_CODE.balance'})
        probe_imd = IMD.create({
            'module': 'suite_cn_statement', 'model': 'account.report.expression',
            'name': 'r46_probe_expr', 'res_id': probe.id})
        try:
            caught = _log_dangling_expressions(self.env)
            self.assertTrue(
                any(lbl == 'r46_probe' for _c, lbl, _f, _m in caught),
                '🔴 R46 哨兵:检测器须抓到注入的悬空表达式;抓不到=检测器空转(本用例本会空绿)')
        finally:
            probe_imd.unlink()
            probe.unlink()
            probe_line.unlink()
        # ② 撤下探针后,真数据无悬空。
        dangling = _log_dangling_expressions(self.env)
        self.assertFalse(
            dangling,
            "generated aggregation expression(s) reference non-existent terms "
            "(report will raise 'Could not expand term' when opened): %s"
            % ', '.join('%s.%s->%s' % (c, l, m) for c, l, _f, m in dangling))


@tagged('post_install', '-at_install')
class TestDisclosureCarrier(TransactionCase):
    """R48-T3 —— 取数口径披露的【载体归属】：中式版式(报送件,税局/审计)撤下页脚,通用披露行
    (对内)保留,独立说明件(客户会计,主动索取)承载,且说明件文案与通用披露行【一字一致】钉住。
    """

    def setUp(self):
        super().setUp()
        self.bs = self.env.ref('l10n_cn_reports.l10n_cn_asbe_bs', raise_if_not_found=False)
        self.line = self.env.ref('suite_cn_statement.cn_asbe_bs_prefix_disclosure',
                                 raise_if_not_found=False)
        if not (self.bs and self.line):
            self.skipTest('ASBE BS / 披露行 未安装')

    # -------------------------------------------------- 判据2 通用披露行仍在
    def test_t3_general_disclosure_line_present(self):
        """判据2：通用路径 account.report.line 披露行仍在、仍挂 ASBE BS（对内受众判断对）。"""
        self.assertTrue(self.line.exists(), '通用披露行不应被撤')
        self.assertEqual(self.line.report_id.id, self.bs.id,
                         '披露行须仍挂在 ASBE 资产负债表上')
        self.assertIn('按科目编码前缀', self.line.name or '', '披露行文案应完整保留')

    # -------------------------------------------------- 判据4 一致性钉（一字一致）
    def test_t3_disclosure_note_matches_general_line(self):
        """判据4（新增一致性钉，R48）：独立说明件文案 == 通用披露行文案，一字一致。
        源同一（说明件从披露行 name 直取），两处再不会漂移（R44 那次「验过没钉住」的补课）。"""
        note = self.bs._cn_disclosure_note()
        self.assertTrue(note, '说明件文案不应为空')
        self.assertEqual(note, self.line.name,
                         '🔴 说明件与通用披露行文案须一字一致，实际已漂移')

    # -------------------------------------------------- 判据3 说明件载体+内容
    def test_t3_note_export_carries_disclosure(self):
        """判据3：独立说明件（工具栏「取数口径说明」）导出内容含该披露文案，贴原样。"""
        action = self.bs.action_export_cn_disclosure_note({})
        self.assertEqual(action.get('type'), 'ir.actions.act_url', '应返回下载动作')
        att_id = int(action['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        body = att.raw.decode('utf-8')
        self.assertIn(self.line.name, body, '说明件正文须含披露行原文')
        self.assertIn('取数口径说明', body, '说明件须标明主题')

    # -------------------------------------------------- 判据1 中式版式不再渲染页脚
    def test_t3_footer_not_rendered_in_cn_layout(self):
        """判据1：中式版式渲染器【不再写 footer_note】——即便 data 里塞了旧页脚串，写入
        单元格命中数=0。用记录式假 worksheet 捕获所有写入文本（xlsx 是 zip 压缩流，字节
        扫描不可靠），与 BS 全链解耦。"""
        class _RecWS:
            def __init__(self):
                self.texts = []
            def write(self, r, c, val, fmt=None):
                self.texts.append(val)
            def merge_range(self, r0, c0, r1, c1, val, fmt=None):
                self.texts.append(val)

        F = {'warn': object(), 'note': object()}
        sentinel = '附注·取数口径：贵司若采用其他编码体系（R48-T3 页脚探针）'
        ws = _RecWS()
        self.bs._cn_xlsx_warning(
            ws, F, {'crossfoot_breaks': [], 'dc_notice': [], 'footer_note': sentinel}, 0, 3)
        self.assertNotIn(sentinel, ws.texts,
                         '🔴 中式版式渲染器仍在写 footer_note（页脚未撤下）')
        # 反哨兵：dc_notice 仍照写 ⇒ 「命中0」不是整段哑掉，而是确实只撤了页脚。
        marker = 'DCNOTICE探针XYZ'
        ws2 = _RecWS()
        self.bs._cn_xlsx_warning(
            ws2, F, {'crossfoot_breaks': [], 'dc_notice': [marker], 'footer_note': ''}, 0, 3)
        self.assertIn(marker, ws2.texts,
                      '反哨兵：dc_notice 应照写——渲染器不是整体哑掉')
