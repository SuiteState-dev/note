# -*- coding: utf-8 -*-
"""R51-T1 —— 中式三栏式明细分类账 验收判据（绝对终态，惯例19；计数类两数分列，惯例17）。

规则锚 = 项目档 v43 §4.5.22（属性制 + 每期两行 + 逐期滚动，唯一归宿）。与总账（R50）同源、
三处形态不同：日期列取分录【实际日期】、有凭证字号列、逐笔明细行 + 逐笔滚动余额。

判据（全部机械可判定）：
  1 逐笔滚动余额：本行 signed == 上一行 signed ± 本行发生额（opening 起链，参与行数 / 不成立）。
  2 本期合计行余额 == 该期最后一条明细行余额（参与期次 / 不成立）。
  3 本期合计行发生额 == 该期全部明细行发生额之和（借/贷各自；参与期次 / 不成立）。
  4 本年累计逐期滚动：本期 ytd(借/贷) == 上一已印期 ytd + 本期合计（参与转移 / 不成立）。
  5 本年累计行余额 == 本期合计行余额（参与对数 / 不相等）。
  6 🔴 跨路径差量检验：明细账本期合计 ↔ 总账同科目同期本期合计（借/贷各自；参与对数 /
    不一致）。⚠️ 名字是「跨路径差量检验」不是「互证」——两表共用上游取数、共模盲区在
    （§4.5.22 七）；覆盖的是两条路径的差量。不一致即【立即停下回报】。
  7 期初余额行形态：凭证字号空、借贷栏空、方向/余额有值、日期 == 范围首日（参与科目 / 不符）。
  8 🔴 日期列取分录实际日期：同期不同日两笔 → 两行日期不同（防「压成期末日」陷阱，§4.5.22 四）。
  9 方向栏：非零取属性、零取平（余额栏同时留空）。必含 ① 反向(3103)；② 平(1122)。
 10 零值留空：应为 0 的金额栏渲染为空、不出 0.00（参与栏位 / 出现 0.00 条数，期望 0）。

惯例30 两半：① 每条断言前置正向哨兵（参与数>0）；② 判据 1/3/4/6/8/10 属校验类，自带
「注入→检出」——在校验边界篡改一格，断言校验确实报出来。
打桩自检（交付要求9 / R46）：核心方法体换 return，对应用例变红——见 stub 自检回报，不在本文件内。
"""
import copy

from odoo.tests import TransactionCase, tagged


@tagged('post_install', '-at_install')
class TestSubsidiaryLedger(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.company.country_id = cls.env.ref('base.cn').id
        cls.currency = cls.company.currency_id
        cls.report = cls.env.ref('suite_cn_statement.cn_subsidiary_ledger')
        cls.handler = cls.env['suite.cn.subsidiary.ledger.report.handler']
        cls.journal = cls.env['account.journal'].search(
            [('type', '=', 'general'), ('company_id', '=', cls.company.id)], limit=1)

        A = cls.env['account.account']

        def acc(code, name, atype):
            return A.create({'name': name, 'code': code, 'account_type': atype})

        # 🔴 夹具按【形态清单】铺，科目个数是结果不是目标（惯例17 第三条第二例）。
        cls.a_cash = acc('1001', 'SL库存现金', 'asset_cash')             # 属性借；多期 + 同期两笔（万能对方腿）
        cls.a_ar = acc('1122', 'SL应收账款', 'asset_current')            # 属性借；【平】(form2) + ≥2 期(薄弱面C)
        cls.a_va = acc('2221.01', 'SL应交增值税', 'liability_current')    # 明细，归 2221（form5 父科目 roll-up）
        cls.a_profit = acc('3103', 'SL本年利润', 'equity')               # 属性贷；反向(form1) + 跨零多行(A) + 拆役(B) + ≥2 期(C)
        cls.a_lti = acc('1601', 'SL长期股权投资', 'asset_non_current')     # 只期初(form3)
        cls.a_inc = acc('6001', 'SL主营业务收入', 'income')               # 损益类结转→平(form4)
        cls.a_dist = acc('3104', 'SL利润分配', 'equity')                 # form4 结转对方

        def post(date, pairs, ref=None):
            move = cls.env['account.move'].create({
                'move_type': 'entry', 'journal_id': cls.journal.id, 'date': date,
                'ref': ref,
                'line_ids': [(0, 0, {'account_id': a.id, 'debit': d, 'credit': c,
                                     'name': nm})
                             for (a, d, c, nm) in pairs],
            })
            move.action_post()
            return move

        # ── 期初（上年，经 to_beginning_of_fiscalyear 进 2026）——form6 跨年期初 ──
        post('2025-12-31', [(cls.a_cash, 1000.0, 0.0, '期初'),
                            (cls.a_va, 0.0, 1000.0, '期初')])
        post('2025-12-31', [(cls.a_ar, 500.0, 0.0, '期初'),
                            (cls.a_profit, 0.0, 500.0, '期初')])      # 3103 期初贷500
        post('2025-12-31', [(cls.a_lti, 700.0, 0.0, '期初'),          # 1601 只期初、本期无发生
                            (cls.a_va, 0.0, 700.0, '期初')])
        # ── 1001 & 2221.01：1 月同月两笔(不同日，判据8) + 3 月 ──
        post('2026-01-10', [(cls.a_cash, 200.0, 0.0, '1.10 收款甲'),
                            (cls.a_va, 0.0, 200.0, '1.10 收款甲')])
        post('2026-01-20', [(cls.a_cash, 100.0, 0.0, '1.20 收款乙'),
                            (cls.a_va, 0.0, 100.0, '1.20 收款乙')])
        post('2026-03-15', [(cls.a_cash, 300.0, 0.0, '3.15 收款丙'),
                            (cls.a_va, 0.0, 300.0, '3.15 收款丙')])
        # ── 1122 平：Feb + May 两期(薄弱面C)，期末归零 ──
        post('2026-02-05', [(cls.a_ar, 0.0, 200.0, '2.5 回款一'),
                            (cls.a_cash, 200.0, 0.0, '2.5 回款一')])
        post('2026-05-10', [(cls.a_ar, 0.0, 300.0, '5.10 回款二'),   # 1122 signed 500-200-300=0 → 平
                            (cls.a_cash, 300.0, 0.0, '5.10 回款二')])
        # ── 3103：四月【同期两笔、余额穿过 0】(薄弱面A) —— 首笔 期首接期初、次笔 同向连续(薄弱面B) ──
        post('2026-04-10', [(cls.a_profit, 200.0, 0.0, '4.10 结转一'),  # signed -500→-300（仍贷余额）
                            (cls.a_cash, 0.0, 200.0, '4.10 结转一')])
        post('2026-04-20', [(cls.a_profit, 400.0, 0.0, '4.20 结转二'),  # signed -300→+100（跨零转借余额=反向）
                            (cls.a_cash, 0.0, 400.0, '4.20 结转二')])
        post('2026-07-15', [(cls.a_profit, 0.0, 50.0, '7.15 调整'),    # 七月：3103 第二期(C)；signed +100→+50 仍反向
                            (cls.a_cash, 50.0, 0.0, '7.15 调整')])
        # ── 6001 + 3104：损益类年末结转→平(form4)。ytd 借==贷、余额归零，与 form2 区分 ──
        post('2026-11-30', [(cls.a_inc, 0.0, 1000.0, '11.30 收入'),
                            (cls.a_cash, 1000.0, 0.0, '11.30 收入')])
        post('2026-12-31', [(cls.a_inc, 1000.0, 0.0, '12.31 结转'),    # 6001 ytd_d 1000==ytd_c 1000, signed 0 → 平
                            (cls.a_dist, 0.0, 1000.0, '12.31 结转')])

        cls.options = cls.report.get_options({
            'selected_variant_id': cls.report.id, 'unfold_all': True,
            'date': {'mode': 'range', 'date_from': '2026-01-01', 'date_to': '2026-12-31'},
        })
        cls.data = cls.handler._cn_sl_compute(cls.report, cls.options)

    # ------------------------------------------------------------- helpers
    def _sec(self, code, data=None):
        data = data or self.data
        for s in data['sections']:
            if s['code'] == code:
                return s
        self.fail('段 %s 不在结果中' % code)

    # 判据1：逐笔滚动余额。opening 起链，逐行 signed 连续。参与=明细行数。
    def _running(self, data):
        n = bad = 0
        for s in data['sections']:
            prev = s['opening']['signed']
            for p in s['periods']:
                for ln in p['lines']:
                    n += 1
                    if abs(ln['signed'] - (prev + ln['debit'] - ln['credit'])) > 0.005:
                        bad += 1
                    prev = ln['signed']
        return n, bad

    # 判据2：本期合计行余额 == 该期最后一条明细行余额。参与=期次。
    def _total_eq_last_line(self, data):
        n = bad = 0
        for s in data['sections']:
            for p in s['periods']:
                if not p['lines']:
                    continue
                n += 1
                if abs(p['current']['signed'] - p['lines'][-1]['signed']) > 0.005:
                    bad += 1
        return n, bad

    # 判据3：本期合计发生额 == Σ明细行发生额（借/贷各自）。参与=期次。
    def _total_eq_sum(self, data):
        n = bad = 0
        for s in data['sections']:
            for p in s['periods']:
                n += 1
                sd = sum(ln['debit'] for ln in p['lines'])
                sc = sum(ln['credit'] for ln in p['lines'])
                if abs(p['current']['debit'] - sd) > 0.005 \
                        or abs(p['current']['credit'] - sc) > 0.005:
                    bad += 1
        return n, bad

    # 判据4：本年累计逐期滚动（借/贷各自）。参与=期次转移。
    def _ytd_rolling(self, data):
        trans = bad = 0
        for s in data['sections']:
            pd = pc = 0.0
            for p in s['periods']:
                trans += 1
                yt, cu = p['ytd'], p['current']
                if abs(yt['debit'] - (pd + cu['debit'])) > 0.005 \
                        or abs(yt['credit'] - (pc + cu['credit'])) > 0.005:
                    bad += 1
                pd, pc = yt['debit'], yt['credit']
        return trans, bad

    # 判据5：本年累计行余额 == 本期合计行余额（同期）。参与=期次对数。
    def _ytd_eq_current(self, data):
        n = bad = 0
        for s in data['sections']:
            for p in s['periods']:
                n += 1
                if p['ytd']['signed'] != p['current']['signed'] \
                        or p['ytd']['balance'] != p['current']['balance']:
                    bad += 1
        return n, bad

    def _gl_current_map(self):
        """总账 handler 每(code, label) 本期合计 (debit, credit)，供判据6 差量对照。"""
        gl = self.env['suite.cn.general.ledger.report.handler']
        glr = self.env.ref('suite_cn_statement.cn_general_ledger')
        gopts = glr.get_options({
            'selected_variant_id': glr.id, 'unfold_all': True,
            'date': {'mode': 'range', 'date_from': '2026-01-01', 'date_to': '2026-12-31'}})
        gdata = gl._cn_gl_compute(glr, gopts)
        m = {}
        for s in gdata['sections']:
            for p in s['periods']:
                m[(s['code'], p['label'])] = (p['current']['debit'], p['current']['credit'])
        return m

    # 判据6：跨路径差量检验（明细账本期合计 vs 总账本期合计）。参与=对数。
    def _cross_path(self, data, gl_map):
        n = bad = 0
        for s in data['sections']:
            for p in s['periods']:
                key = (s['code'], p['label'])
                if key not in gl_map:
                    continue
                n += 1
                gd, gc = gl_map[key]
                if abs(p['current']['debit'] - gd) > 0.005 \
                        or abs(p['current']['credit'] - gc) > 0.005:
                    bad += 1
        return n, bad

    # 判据7：期初行形态。参与=期初非零段。
    def _opening_form(self, data):
        n = bad = 0
        for s in data['sections']:
            o = s['opening']
            if o['balance'] is None:          # 期初为平——不入参与集
                continue
            n += 1
            if not o['direction'] or o['balance'] is None \
                    or s['open_date'] != '2026-01-01':
                bad += 1
        return n, bad

    # =============================================================== 判据1
    def test_j1_running_balance(self):
        n, bad = self._running(self.data)
        self.assertGreater(n, 0, '哨兵：参与明细行数须>0')
        self.assertEqual(bad, 0, '判据1：逐笔滚动余额连续（参与 %d / 不成立 %d）' % (n, bad))

    def test_j1_running_inject_detect(self):
        tampered = copy.deepcopy(self.data)
        sec = self._sec('1001', tampered)
        sec['periods'][0]['lines'][1]['signed'] += 999.0
        n, bad = self._running(tampered)
        self.assertGreaterEqual(bad, 1, '🔴 逐笔滚动校验未检出注入的断裂')

    # =============================================================== 判据2
    def test_j2_total_equals_last_line(self):
        n, bad = self._total_eq_last_line(self.data)
        self.assertGreater(n, 0, '哨兵：参与期次须>0')
        self.assertEqual(bad, 0, '判据2：本期合计余额==该期末行余额（参与 %d / 不成立 %d）' % (n, bad))

    # =============================================================== 判据3
    def test_j3_total_equals_line_sum(self):
        n, bad = self._total_eq_sum(self.data)
        self.assertGreater(n, 0, '哨兵：参与期次须>0')
        self.assertEqual(bad, 0, '判据3：本期合计发生额==Σ明细行（参与 %d / 不成立 %d）' % (n, bad))
        # 1001 一月两笔合计 借 300
        jan = self._sec('1001')['periods'][0]
        self.assertEqual(len(jan['lines']), 2, '1001 一月应有 2 笔明细')
        self.assertAlmostEqual(jan['current']['debit'], 300.0, places=2)

    def test_j3_sum_inject_detect(self):
        tampered = copy.deepcopy(self.data)
        self._sec('1001', tampered)['periods'][0]['current']['debit'] += 50.0
        n, bad = self._total_eq_sum(tampered)
        self.assertGreaterEqual(bad, 1, '🔴 本期合计=Σ明细 校验未检出注入')

    # =============================================================== 判据4
    def test_j4_ytd_rolling(self):
        trans, bad = self._ytd_rolling(self.data)
        self.assertGreater(trans, 0, '哨兵：参与转移数须>0')
        self.assertEqual(bad, 0, '判据4：本年累计逐期滚动（参与 %d / 不成立 %d）' % (trans, bad))
        self.assertGreaterEqual(len(self._sec('1001')['periods']), 3, '1001 应有≥3 个期')

    def test_j4_ytd_inject_detect(self):
        tampered = copy.deepcopy(self.data)
        self._sec('1001', tampered)['periods'][2]['ytd']['debit'] += 77.0
        trans, bad = self._ytd_rolling(tampered)
        self.assertGreaterEqual(bad, 1, '🔴 逐期滚动校验未检出注入的累计断裂')

    # =============================================================== 判据5
    def test_j5_ytd_equals_current_balance(self):
        n, bad = self._ytd_eq_current(self.data)
        self.assertGreater(n, 0, '哨兵：参与期次对数须>0')
        self.assertEqual(bad, 0, '判据5：本年累计行余额==本期合计行余额（参与 %d / 不相等 %d）' % (n, bad))

    # =============================================================== 判据6
    def test_j6_cross_path_diff_check(self):
        gl_map = self._gl_current_map()
        n, bad = self._cross_path(self.data, gl_map)
        self.assertGreater(n, 0, '哨兵：跨路径对数须>0')
        self.assertEqual(bad, 0,
                         '🔴 判据6：明细账本期合计 与总账不一致（参与 %d / 不一致 %d）'
                         '——跨路径差量检验，不一致即立即停下' % (n, bad))

    def test_j6_cross_path_inject_detect(self):
        gl_map = self._gl_current_map()
        tampered = copy.deepcopy(self.data)
        self._sec('1001', tampered)['periods'][0]['current']['debit'] += 123.0
        n, bad = self._cross_path(tampered, gl_map)
        self.assertGreaterEqual(bad, 1, '🔴 跨路径差量检验未检出注入的两表不一致')

    # =============================================================== 判据7
    def test_j7_opening_row_form(self):
        n, bad = self._opening_form(self.data)
        self.assertGreater(n, 0, '哨兵：期初非零段数须>0')
        self.assertEqual(bad, 0, '判据7：期初行形态（参与 %d / 不符 %d）' % (n, bad))
        # 期初行在渲染层：凭证字号空、借贷空、日期=范围首日
        rows = self.report._cn_sl_prepare(self.options)['sections']
        s1001 = next(s for s in rows if s['code'] == '1001')
        op_row = s1001['rows'][0]
        self.assertEqual(op_row['summary'], '期初余额')
        self.assertEqual(op_row['voucher'], '', '期初行凭证字号须空')
        self.assertIsNone(op_row['debit'], '期初行借方栏须空')
        self.assertIsNone(op_row['credit'], '期初行贷方栏须空')
        self.assertEqual(op_row['date'], '2026-01-01', '期初行日期须=范围首日')

    # =============================================================== 判据8
    def test_j8_date_is_actual_entry_date(self):
        """同期不同日两笔 → 两行日期不同，且是实际日期，非期末日（§4.5.22 四陷阱）。"""
        jan = self._sec('1001')['periods'][0]
        dates = [ln['date'] for ln in jan['lines']]
        self.assertEqual(len(dates), 2, '一月应有两笔')
        self.assertEqual(sorted(dates), ['2026-01-10', '2026-01-20'],
                         '🔴 日期须取分录实际日期')
        squashed = sum(1 for d in dates if d == '2026-01-31')
        self.assertEqual(squashed, 0, '🔴 日期被压成期末日的条数须【确认为 0】')

    def test_j8_squash_inject_detect(self):
        """注入→检出：把两行日期都改成期末日，断言检测器报出被压。"""
        jan = copy.deepcopy(self._sec('1001'))['periods'][0]
        for ln in jan['lines']:
            ln['date'] = '2026-01-31'
        dates = [ln['date'] for ln in jan['lines']]
        squashed = sum(1 for d in dates if d == '2026-01-31')
        self.assertGreaterEqual(squashed, 1, '🔴 期末日压平检测器未检出注入')

    # =============================================================== 判据9
    def test_j9_direction_reversal_and_ping(self):
        """① 反向：3103 属性贷、期末借方余额 ⇒ 方向贷、余额负。② 平：1122 期末零 ⇒ 平、余额空。"""
        profit = self._sec('3103')
        self.assertEqual(profit['total']['direction'], '贷',
                         '🔴 反向：属性贷科目实际借方余额，方向仍取属性「贷」')
        self.assertLess(profit['total']['balance'], 0, '🔴 反向：余额须为负')
        ar = self._sec('1122')
        self.assertEqual(ar['total']['direction'], '平', '🔴 平：期末零余额方向须为「平」')
        self.assertIsNone(ar['total']['balance'], '🔴 平：余额栏须留空（None）')

    def test_j9_dir_bal_reuse_gl(self):
        """复用总账 _cn_gl_dir_bal（§三 复用四条）：单测钉住属性定向 + 符号 + 平。"""
        gl = self.env['suite.cn.general.ledger.report.handler']
        f, cur = gl._cn_gl_dir_bal, self.currency
        self.assertEqual(f(300.0, 'credit', cur), ('贷', -300.0))   # 属性贷/实际借 ⇒ 负
        self.assertEqual(f(0.0, 'debit', cur), ('平', None))

    # =============================================================== 判据10
    def _zero_cells_in_data(self, data):
        n = 0
        for s in data['sections']:
            for p in s['periods']:
                for blk in ('current', 'ytd'):
                    for k in ('debit', 'credit'):
                        if abs(p[blk][k]) < 1e-9:
                            n += 1
                for ln in p['lines']:
                    for k in ('debit', 'credit'):
                        if abs(ln[k]) < 1e-9:
                            n += 1
        return n

    def _render_and_capture(self, data_override=None):
        data = data_override or self.report._cn_sl_prepare(self.options)
        rec = _RecWS()
        keys = ['title', 'meta', 'meta_r', 'hdr', 'acct', 'no', 'name', 'name_b',
                'num', 'num_b', 'empty', 'sign', 'note']
        F = {k: object() for k in keys}
        self.report._cn_xlsx_sl(rec, F, data)
        return rec

    def test_j10_zero_blank_no_double_zero(self):
        participate = self._zero_cells_in_data(self.data)
        self.assertGreater(participate, 0, '哨兵：应为 0 的发生额栏位须>0')
        rec = self._render_and_capture()
        leaked = sum(1 for v in rec.numbers if abs(v) < 1e-9)
        self.assertEqual(leaked, 0,
                         '判据10：零值须渲染为空（参与 %d / 出现 0.00 条数 %d，期望 0）'
                         % (participate, leaked))

    def test_j10_zero_leak_inject_detect(self):
        rec = _RecWS()
        rec.write_number(0, 0, 0.0, None)
        leaked = sum(1 for v in rec.numbers if abs(v) < 1e-9)
        self.assertGreaterEqual(leaked, 1, '🔴 0.00 泄漏检测器未检出注入的零值')

    # =============================================================== 结构 / rollup
    def test_only_level1_rollup(self):
        self.assertEqual(self.data['detail_leak'], 0,
                         '明细科目作独立段的条数须【确认为 0】（roll-up 到一级）')
        codes = [s['code'] for s in self.data['sections']]
        self.assertIn('2221', codes, '一级键 2221 应成段')
        self.assertNotIn('2221.01', codes, '明细 2221.01 不应作独立段')

    # =============================================================== 渲染两路径
    def test_screen_lines_structure(self):
        lines = self.report._get_lines(self.options)
        self.assertGreater(len(lines), 0, '哨兵：屏幕行须>0')
        names = [l.get('name') for l in lines]
        self.assertTrue(any(n and n.startswith('1001 ') for n in names), '应有 1001 段头')
        self.assertTrue(any('期初余额' in (n or '') for n in names))
        self.assertTrue(any('本期合计' in (n or '') for n in names), '应有本期合计行')
        self.assertTrue(any('本年累计' in (n or '') for n in names))
        self.assertTrue(any('收款甲' in (n or '') for n in names), '逐笔摘要须落 name')
        self.assertFalse(any((n or '').startswith('Total ') for n in names),
                         '🔴 不应出现框架 totals_below_sections 噪声合计行')

    def test_cn_xlsx_sl_render(self):
        data = self.report._cn_sl_prepare(self.options)
        rec = self._render_and_capture(data)
        joined = '\n'.join(t for t in rec.texts if isinstance(t, str))
        self.assertIn('三栏式明细分类账', joined, '表名须在')
        self.assertIn('科目：1001', joined, '科目 banner 须在（导出件里科目在表头）')
        self.assertIn('凭证字号', joined, '凭证字号列头须在')
        self.assertIn('2026-01-10', joined, '逐笔行实际日期须在（判据8）')
        self.assertIn('本期合计', joined)
        self.assertIn('本年累计', joined)
        self.assertIn('贷', rec.texts, '方向标签须写入（2221/3103 属性贷）')
        self.assertIn('口径一致', joined, '口径提示（与 TB/总账一致）须在表内')

    def test_t2_footer_company_and_pagenum(self):
        """R52-T2：每页页脚含【公司名】+【连续页码】两者并存（导出件锚：公司名在页脚非表头）。"""
        data = self.report._cn_sl_prepare(self.options)
        rec = self._render_and_capture(data)
        self.assertIsNotNone(rec.footer, '页脚须被设置')
        self.assertIn('公司名称：', rec.footer, '🔴 页脚须含公司名（R52-T2 订正 R51 表头误判）')
        self.assertIn(data['company'], rec.footer, '页脚公司名须为本公司名')
        self.assertIn('&P', rec.footer, '§58 连续页码（当前页 &P）须仍在')
        self.assertIn('&N', rec.footer, '§58 连续页码（总页 &N）须仍在')

    def test_t2_header_no_company(self):
        """R53-T2：明细账【表头】不出公司名（taxpayer_name 置空）——载体落点闭合（kingdee v6：
        明细账公司名在页脚、总账才在表头）。表头区渲染文本不得出现公司名。"""
        data = self.report._cn_sl_prepare(self.options)
        self.assertEqual(data['taxpayer_name'], '', '🔴 表头纳税人名称须置空（公司名不入表头）')
        rec = self._render_and_capture(data)
        header_texts = '\n'.join(t for t in rec.texts if isinstance(t, str))
        self.assertNotIn(data['company'], header_texts, '🔴 表头区不得出现公司名')

    def test_t2_openpyxl_real_file(self):
        """R53-T2 判据：openpyxl 读【真件】——表头单元格无公司名、页脚串含公司名 + 页码。
        （直接以 xlsxwriter 出件再读，避开 export_to_cn_sl_xlsx 强制 lang=zh_CN；版式代码同一条。）"""
        import io as _io
        try:
            import openpyxl  # noqa: PLC0415
            import xlsxwriter  # noqa: PLC0415
        except ImportError:
            self.skipTest('openpyxl/xlsxwriter 不可用')
        data = self.report._cn_sl_prepare(self.options)
        buf = _io.BytesIO()
        wb0 = xlsxwriter.Workbook(buf, {'in_memory': True})
        ws0 = wb0.add_worksheet('SL')
        keys = ['title', 'meta', 'meta_r', 'hdr', 'acct', 'no', 'name', 'name_b',
                'num', 'num_b', 'empty', 'sign', 'note']
        F = {k: wb0.add_format({}) for k in keys}
        self.report._cn_xlsx_sl(ws0, F, data)
        wb0.close()
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        # 表头区（前 5 行 = 表名/表号/纳税人/所属期/列头）不得出现公司名
        header_cells = []
        for row in ws.iter_rows(min_row=1, max_row=5):
            for c in row:
                if c.value is not None:
                    header_cells.append(str(c.value))
        header_blob = '\n'.join(header_cells)
        self.assertNotIn(data['company'], header_blob,
                         '🔴 表头区无公司名（原样表头：%s）' % header_cells)
        # 页脚（openpyxl 读 oddFooter）含公司名 + 连续页码
        of = ws.oddFooter
        footer_str = '%s|%s|%s' % (of.left.text or '', of.center.text or '', of.right.text or '')
        self.assertIn('公司名称：', footer_str, '🔴 页脚含公司名（原样页脚：%r）' % footer_str)
        self.assertIn(data['company'], footer_str)
        self.assertIn('&P', footer_str, '连续页码 &P 须在')
        self.assertIn('&N', footer_str, '连续页码 &N 须在')


    # =============================== 薄弱面A：反向科目多行滚动 + 中途跨零
    def _dir_by_attr(self, data):
        """每条非零明细行的方向须 == 科目属性（借/贷），余额跨零也不翻向。参与=非零明细行。"""
        want = {'debit': '借', 'credit': '贷'}
        n = bad = 0
        for s in data['sections']:
            w = want.get(s['attr'])
            for p in s['periods']:
                for ln in p['lines']:
                    if abs(ln['signed']) < 1e-9:
                        continue
                    n += 1
                    if ln['direction'] != w:
                        bad += 1
        return n, bad

    def test_a_dir_by_attr_midstream_crossing(self):
        """薄弱面A：3103 四月同期两笔、余额逐笔滚动【穿过 0】；属性制方向逐行不翻向。"""
        n, bad = self._dir_by_attr(self.data)
        self.assertGreater(n, 0, '哨兵：非零明细行须>0')
        self.assertEqual(bad, 0, '薄弱面A：方向逐行==科目属性（参与 %d / 翻向 %d）' % (n, bad))
        apr = self._sec('3103')['periods'][0]
        self.assertEqual(len(apr['lines']), 2, '3103 四月须两笔（多行滚动 + 跨零）')
        l0, l1 = apr['lines']
        self.assertLess(l0['signed'], 0, '跨零前：首笔后 signed<0（仍贷方余额）')
        self.assertGreater(l1['signed'], 0, '🔴 跨零后：次笔后 signed>0（转借方余额=反向）')
        self.assertEqual([l0['direction'], l1['direction']], ['贷', '贷'],
                         '🔴 属性制：余额跨零转正后方向仍取属性「贷」，不翻「借」')
        self.assertGreater(l0['balance'], 0, '贷余额：balance = -signed > 0')
        self.assertLess(l1['balance'], 0, '借余额(反向)：balance = -signed < 0')

    def test_a_dir_inject_detect(self):
        """注入→检出：模拟「余额符号制」错误实现——跨零后把方向翻成「借」。"""
        tampered = copy.deepcopy(self.data)
        self._sec('3103', tampered)['periods'][0]['lines'][1]['direction'] = '借'
        n, bad = self._dir_by_attr(tampered)
        self.assertGreaterEqual(bad, 1, '🔴 方向==属性校验未检出注入的跨零翻向')

    # =============================== 薄弱面B：期首接期初 / 同向连续 拆两行
    def test_b_role_split_two_rows(self):
        """薄弱面B：3103 四月——行0 只承担【期首接期初】、行1 只承担【同向连续】，两个不同的行。"""
        sec = self._sec('3103')
        apr = sec['periods'][0]
        self.assertGreaterEqual(len(apr['lines']), 2, '需≥2行以分离两种角色')
        l0, l1 = apr['lines'][0], apr['lines'][1]
        # 期首接期初：行0 滚动自【期初余额行】
        self.assertAlmostEqual(
            l0['signed'], sec['opening']['signed'] + l0['debit'] - l0['credit'], places=2,
            msg='期首接期初：行0 signed 承接期初余额行')
        # 同向连续：行1 滚动自【上一明细行(行0)】
        self.assertAlmostEqual(
            l1['signed'], l0['signed'] + l1['debit'] - l1['credit'], places=2,
            msg='同向连续：行1 signed 承接行0')
        # 🔴 两角色落在两个不同的行
        self.assertIsNot(l0, l1, '两角色须为两个不同的行对象')
        # 行1 不从期初直接滚出——证明它只承担「同向连续」、不兼「期首接期初」
        self.assertNotAlmostEqual(
            l1['signed'], sec['opening']['signed'] + l1['debit'] - l1['credit'], places=2,
            msg='🔴 行1 不应从期初直接滚出（否则又是单行双役）')

    # =============================== 形态覆盖（六类全有）
    def test_t1_form_coverage(self):
        """SL 夹具六类形态逐类核【补后全部有】——缺哪类补哪类，科目个数是结果不是目标。"""
        forms = self._detect_forms(self.data)
        for key, label in [('f1_reverse', '反向'), ('f2_ping', '平'),
                           ('f3_only_opening', '只期初'), ('f4_carry', '损益结转→平'),
                           ('f5_rollup', '父科目 roll-up'), ('f6_cross_year_open', '跨年期初')]:
            self.assertTrue(forms[key],
                            'SL 夹具缺形态【%s】（承载科目 %s）' % (label, forms[key]))

    # ---------------------------------------------------- 形态检测（六类）
    def _detect_forms(self, data):
        """按段实测六类形态，返回 {form: [承载科目 code...]}。科目个数是结果不是目标。"""
        secs = data['sections']
        f1 = [s['code'] for s in secs
              if s['total']['balance'] is not None and s['total']['balance'] < 0]
        f2 = [s['code'] for s in secs if s['total']['direction'] == '平'
              and not (s['total']['debit'] > 1e-9 and s['total']['credit'] > 1e-9)]
        f3 = [s['code'] for s in secs if not s['periods']]
        f4 = [s['code'] for s in secs if s['total']['direction'] == '平'
              and s['total']['debit'] > 1e-9 and s['total']['credit'] > 1e-9]
        f5 = [s['code'] for s in secs
              if self.env['account.account'].search_count(
                  [('code', '=like', s['code'] + '.%')])]
        f6 = [s['code'] for s in secs
              if abs(s['opening']['signed']) > 1e-9]
        return {'f1_reverse': f1, 'f2_ping': f2, 'f3_only_opening': f3,
                'f4_carry': f4, 'f5_rollup': f5, 'f6_cross_year_open': f6}

class _RecWS:
    """记录式假 worksheet：捕获文本与数字写入，供渲染/零值泄漏断言。"""
    def __init__(self):
        self.texts = []
        self.numbers = []
        self.footer = None

    def write(self, r, c, val, fmt=None):
        self.texts.append(val)

    def write_number(self, r, c, val, fmt=None):
        self.numbers.append(val)

    def write_blank(self, r, c, val, fmt=None):
        pass

    def merge_range(self, r0, c0, r1, c1, val, fmt=None):
        self.texts.append(val)

    def set_column(self, *a, **k):
        pass

    def set_row(self, *a, **k):
        pass

    def set_footer(self, s=None, *a, **k):
        self.footer = s
